# data_entry_v2_dynamic_excel_with_validation.py
# v2+: Dynamic Excel-driven Tkinter data-entry app with adaptive validation
#
# Save as: data_entry_v2_dynamic_excel_with_validation.py
# Requirements:
# pip install ttkbootstrap openpyxl

import os
import re
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from ttkbootstrap import Window, Style
from openpyxl import load_workbook, Workbook
from datetime import datetime

APP_TITLE = "Data Entry v2+ — Dynamic Excel Companion (with Validation)"


# -------------------------
# Validation helpers
# -------------------------
EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")

def try_parse_date(value: str):
    """Try multiple common date formats. Return date object if successful, else None."""
    if value is None:
        return None
    s = value.strip()
    if s == "":
        return None
    formats = ["%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y", "%Y/%m/%d"]
    for fmt in formats:
        try:
            return datetime.strptime(s, fmt).date()
        except Exception:
            continue
    # try ISO fallback
    try:
        return datetime.fromisoformat(s).date()
    except Exception:
        return None

def is_numeric(value: str):
    if value is None:
        return False
    s = str(value).strip()
    if s == "":
        return False
    try:
        float(s.replace(",", ""))  # allow commas in thousands
        return True
    except Exception:
        return False

def normalize_numeric(value: str):
    return float(str(value).strip().replace(",", ""))


# -------------------------
# Main app
# -------------------------
class DynamicExcelApp:
    def __init__(self, root: Window):
        self.root = root
        self.root.title(APP_TITLE)
        self.root.geometry("1100x720")

        # State
        self.workbook = None
        self.filepath = None
        self.active_sheet_name = None
        self.headers = []
        self.input_entries = []  # list of tk.Entry widgets matching headers
        self.unsaved_changes = False

        # Inferred validation rules per column: list of dicts with keys: type ('text','numeric','date','email'), required(bool)
        self.validation_rules = []

        # UI elements
        self._create_menu()
        self._create_toolbar()
        self._create_top_frame()
        self._create_bottom_frame()
        self._bind_events()

        # Start with a clean app — prompt user to open file
        self._prompt_open_file_on_startup()

    # -----------------------
    # UI Construction
    # -----------------------
    def _create_menu(self):
        menubar = tk.Menu(self.root)
        file_menu = tk.Menu(menubar, tearoff=0)
        file_menu.add_command(label="Open...", command=self.open_file)
        file_menu.add_command(label="Save", command=self.save_file)
        file_menu.add_command(label="Save As...", command=self.save_file_as)
        file_menu.add_separator()
        file_menu.add_command(label="Exit", command=self.on_close)
        menubar.add_cascade(label="File", menu=file_menu)

        help_menu = tk.Menu(menubar, tearoff=0)
        help_menu.add_command(label="About", command=self._show_about)
        menubar.add_cascade(label="Help", menu=help_menu)

        self.root.config(menu=menubar)

    def _create_toolbar(self):
        toolbar = ttk.Frame(self.root)
        toolbar.pack(side=tk.TOP, fill=tk.X, padx=6, pady=6)

        open_btn = ttk.Button(toolbar, text="Open File", command=self.open_file)
        open_btn.pack(side=tk.LEFT, padx=(0, 6))

        save_btn = ttk.Button(toolbar, text="Save", command=self.save_file)
        save_btn.pack(side=tk.LEFT, padx=(0, 6))

        self.add_button = ttk.Button(toolbar, text="Add Row", command=self.add_row_from_inputs)
        self.add_button.pack(side=tk.LEFT, padx=(0, 6))

        ttk.Label(toolbar, text="Sheet:").pack(side=tk.LEFT, padx=(12, 4))
        self.sheet_combo = ttk.Combobox(toolbar, state="readonly", width=30)
        self.sheet_combo.pack(side=tk.LEFT)
        self.sheet_combo.bind("<<ComboboxSelected>>", self.on_sheet_change)

        # Spacer
        spacer = ttk.Label(toolbar, text="")
        spacer.pack(side=tk.LEFT, expand=True)

        # Theme toggler using ttkbootstrap's Style
        ttk.Label(toolbar, text="Theme:").pack(side=tk.LEFT, padx=(6, 4))
        self.theme_combo = ttk.Combobox(toolbar, values=Style().theme_names(), state="readonly", width=15)
        self.theme_combo.set(Style().theme_use())
        self.theme_combo.bind("<<ComboboxSelected>>", self.on_theme_change)
        self.theme_combo.pack(side=tk.LEFT, padx=(0, 6))

        # Status label
        self.status_var = tk.StringVar(value="No file opened.")
        self.status_label = ttk.Label(toolbar, textvariable=self.status_var)
        self.status_label.pack(side=tk.RIGHT)

    def _create_top_frame(self):
        # top frame will contain dynamic input fields (headers -> entries)
        self.top_frame = ttk.Frame(self.root)
        self.top_frame.pack(side=tk.TOP, fill=tk.X, padx=10, pady=(6, 0))

        # We'll put inputs into a scrollable frame in case there are many columns
        self.input_canvas = tk.Canvas(self.top_frame, height=140)
        self.input_canvas.pack(side=tk.TOP, fill=tk.X, expand=True)

        self.input_scrollbar = ttk.Scrollbar(self.top_frame, orient=tk.HORIZONTAL, command=self.input_canvas.xview)
        self.input_scrollbar.pack(side=tk.TOP, fill=tk.X)
        self.input_canvas.configure(xscrollcommand=self.input_scrollbar.set)

        self.inputs_inner = ttk.Frame(self.input_canvas)
        self.input_canvas.create_window((0, 0), window=self.inputs_inner, anchor="nw")
        self.inputs_inner.bind("<Configure>", lambda e: self.input_canvas.configure(scrollregion=self.input_canvas.bbox("all")))

    def _create_bottom_frame(self):
        # bottom frame will contain the treeview of sheet data
        bottom_frame = ttk.Frame(self.root)
        bottom_frame.pack(side=tk.TOP, fill=tk.BOTH, expand=True, padx=10, pady=10)

        self.tree = ttk.Treeview(bottom_frame, show="headings")
        self.tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        vsb = ttk.Scrollbar(bottom_frame, orient="vertical", command=self.tree.yview)
        vsb.pack(side=tk.LEFT, fill=tk.Y)
        self.tree.configure(yscrollcommand=vsb.set)

        hsb = ttk.Scrollbar(bottom_frame, orient="horizontal", command=self.tree.xview)
        hsb.pack(side=tk.BOTTOM, fill=tk.X)
        self.tree.configure(xscrollcommand=hsb.set)

    def _bind_events(self):
        # Save prompt on close
        self.root.protocol("WM_DELETE_WINDOW", self.on_close)

    # -----------------------
    # File and Sheet Handling
    # -----------------------
    def _prompt_open_file_on_startup(self):
        # Clean workspace: prompt for file
        answer = messagebox.askyesno("Open file", "Do you want to open an existing .xlsx file to work on?")
        if answer:
            self.open_file()
        else:
            self.status_var.set("Ready. Use File -> Open to open an .xlsx file.")

    def open_file(self):
        filetypes = [("Excel files", "*.xlsx")]
        path = filedialog.askopenfilename(title="Open Excel file", filetypes=filetypes)
        if not path:
            return

        try:
            wb = load_workbook(path)
        except Exception as e:
            messagebox.showerror("Error", f"Failed to open workbook:\n{e}")
            return

        self.workbook = wb
        self.filepath = path
        # default to active sheet
        self.active_sheet_name = self.workbook.active.title
        self._populate_sheet_selector()
        self._load_active_sheet()
        self.unsaved_changes = False
        self._update_status(f"Opened: {os.path.basename(path)}")

    def _populate_sheet_selector(self):
        if not self.workbook:
            self.sheet_combo["values"] = []
            return
        names = self.workbook.sheetnames
        self.sheet_combo["values"] = names
        if self.active_sheet_name in names:
            self.sheet_combo.set(self.active_sheet_name)
        else:
            self.sheet_combo.set(names[0])
            self.active_sheet_name = names[0]

    def on_sheet_change(self, event=None):
        if not self.workbook:
            return
        new_sheet = self.sheet_combo.get()
        if new_sheet == self.active_sheet_name:
            return

        if self.unsaved_changes:
            res = messagebox.askyesnocancel("Unsaved changes", "You have unsaved changes. Save before switching sheets?")
            if res is None:
                # Cancel switching
                self.sheet_combo.set(self.active_sheet_name)
                return
            if res:  # Yes -> save
                if not self.save_file():
                    # Save failed or cancelled; cancel switching
                    self.sheet_combo.set(self.active_sheet_name)
                    return

        self.active_sheet_name = new_sheet
        self._load_active_sheet()

    def _load_active_sheet(self):
        """Read headers and rows from active sheet, rebuild inputs and treeview."""
        if not self.workbook or not self.active_sheet_name:
            return

        sheet = self.workbook[self.active_sheet_name]
        # Read headers: find first non-empty row (usually row 1). We treat the very first row with any non-empty cells as headers.
        headers = []
        header_row_idx = None
        for r in sheet.iter_rows(min_row=1, max_row=5):  # check first few rows for header row
            values = [cell.value for cell in r]
            if any(v is not None and str(v).strip() != "" for v in values):
                header_row_idx = r[0].row
                headers = []
                for cell in r:
                    # Stop at trailing blank cells? We'll keep blanks but give them fallback names
                    val = cell.value
                    headers.append(str(val).strip() if val is not None and str(val).strip() != "" else None)
                break

        if not headers:
            # If still no headers, create default Column1..N based on max_column
            max_col = sheet.max_column or 1
            headers = [None] * max_col
            header_row_idx = 1

        # Normalize headers: fill None with "Column X"
        normalized = []
        for i, h in enumerate(headers):
            if h:
                normalized.append(h)
            else:
                normalized.append(f"Column {i+1}")
        self.headers = normalized

        # Infer validation rules
        self.validation_rules = self._infer_validation_rules(self.headers)

        # Build inputs and treeview
        self._build_input_fields(self.headers)
        self._load_treeview_rows(sheet, header_row_idx)

    # -----------------------
    # Validation rule inference
    # -----------------------
    def _infer_validation_rules(self, headers):
        """
        Heuristic inference of validation rules from header names.
        Returns a list of dicts: {'type': 'text'|'numeric'|'date'|'email', 'required': bool}
        """
        rules = []
        numeric_keywords = ("qty", "quantity", "amount", "price", "total", "number", "age", "count")
        for h in headers:
            h_lower = h.lower() if h else ""
            rule = {"type": "text", "required": True}
            # optional indicator
            if "optional" in h_lower or "[optional]" in h_lower or "(optional)" in h_lower:
                rule["required"] = False
            # date
            if "date" in h_lower or "dob" in h_lower or "birth" in h_lower:
                rule["type"] = "date"
            # numeric
            elif any(k in h_lower for k in numeric_keywords):
                rule["type"] = "numeric"
            # email
            elif "email" in h_lower:
                rule["type"] = "email"
            # else keep text
            rules.append(rule)
        return rules

    # -----------------------
    # Inputs builder & validation
    # -----------------------
    def _clear_inputs_area(self):
        for w in self.inputs_inner.winfo_children():
            w.destroy()
        self.input_entries.clear()

    def _build_input_fields(self, headers):
        """Create a horizontal layout of label+entry for each header."""
        self._clear_inputs_area()

        # Keep everything left to right inside inputs_inner using grid
        for idx, header in enumerate(headers):
            col_frame = ttk.Frame(self.inputs_inner)
            col_frame.grid(row=0, column=idx, padx=6, pady=4)
            lbl = ttk.Label(col_frame, text=header, width=20, anchor="center")
            lbl.pack(side=tk.TOP, fill=tk.X)
            # Use tk.Entry for easy bg color changes on validation
            ent = tk.Entry(col_frame, width=20)
            ent.pack(side=tk.TOP, pady=(6, 0))
            # Pressing Enter in last entry triggers add row; else move to next
            ent.bind("<Return>", lambda e, i=idx: self._on_enter_pressed(e, i))
            self.input_entries.append(ent)

        # Add width fudge: update the canvas scroll region after a small delay
        self.root.after(100, lambda: self.input_canvas.configure(scrollregion=self.input_canvas.bbox("all")))

    def _on_enter_pressed(self, event, idx):
        # If Enter pressed in last field, add row; else focus next.
        if idx == len(self.input_entries) - 1:
            self.add_row_from_inputs()
        else:
            self.input_entries[idx + 1].focus_set()

    def validate_inputs(self):
        """
        Validate all input entries based on inferred self.validation_rules.
        Returns (is_valid: bool, messages: list, normalized_values: list)
        normalized_values is suitable for writing into Excel (numbers, ISO date strings, or text).
        """
        values = [entry.get().strip() for entry in self.input_entries]
        normalized = []
        messages = []
        is_valid = True

        for i, (val, rule) in enumerate(zip(values, self.validation_rules)):
            col_name = rule.get("name", f"Column {i+1}")
            val_type = rule.get("type", "text")
            required = rule.get("required", False)

            # Required check
            if required and not val:
                is_valid = False
                messages.append(f"{col_name}: This field is required.")
                normalized.append(None)
                continue

            # --- Text normalization ---
            if val_type == "text":
                if val:
                    # Collapse multiple spaces, trim, and normalize capitalization if desired
                    cleaned = re.sub(r"\s+", " ", val.strip())
                    normalized.append(cleaned)
                else:
                    normalized.append(None)

            # --- Numeric validation ---
            elif val_type == "numeric":
                if val:
                    try:
                        normalized.append(float(val))
                    except ValueError:
                        is_valid = False
                        messages.append(f"{col_name}: Invalid numeric value.")
                        normalized.append(None)
                else:
                    normalized.append(None)

            # --- Date validation ---
            elif val_type == "date":
                if val:
                    try:
                        normalized.append(datetime.fromisoformat(val).date().isoformat())
                    except Exception:
                        is_valid = False
                        messages.append(f"{col_name}: Invalid date format (use YYYY-MM-DD).")
                        normalized.append(val)
                else:
                    normalized.append(None)

            else:
                normalized.append(val)

        return is_valid, messages, normalized

    def clear_input_entries(self):
        for ent in self.input_entries:
            ent.delete(0, tk.END)
            # Reset background
            ent.config(bg="white")

    # -----------------------
    # Treeview population
    # -----------------------
    def _clear_treeview(self):
        for col in self.tree["columns"]:
            self.tree.heading(col, text="")
        self.tree.delete(*self.tree.get_children())
        self.tree["columns"] = ()

    def _load_treeview_rows(self, sheet, header_row_idx):
        self._clear_treeview()
        # Treeview columns -> headers
        cols = [f"c{i}" for i in range(len(self.headers))]
        self.tree["columns"] = cols
        for i, h in enumerate(self.headers):
            self.tree.heading(cols[i], text=h, anchor=tk.W)
            # set column width heuristically
            self.tree.column(cols[i], width=160, anchor=tk.W)

        # Read rows starting after header_row_idx
        start_row = header_row_idx + 1
        rows = []
        for r in sheet.iter_rows(min_row=start_row, max_row=sheet.max_row, max_col=len(self.headers)):
            rowvals = [ (cell.value if cell.value is not None else "") for cell in r ]
            # If the row is entirely empty, skip
            if all((v == "" or v is None) for v in rowvals):
                continue
            rows.append(rowvals)

        for row in rows:
            # Ensure row length matches headers
            row_extended = list(row) + [""] * (len(self.headers) - len(row))
            self.tree.insert("", tk.END, values=row_extended)

    # -----------------------
    # Add / Edit Row + Save
    # -----------------------
    def add_row_from_inputs(self):
        """Append a new row after validating input fields."""
        if not self.workbook or not self.active_sheet_name:
            messagebox.showwarning("No file", "Open an .xlsx file first.")
            return

        # Validate inputs
        is_valid, messages, normalized = self.validate_inputs()
        if not is_valid:
            messagebox.showwarning("Validation failed", "\n".join(messages))
            return

        sheet = self.workbook[self.active_sheet_name]
        append_row_idx = sheet.max_row + 1

        # Write values into workbook
        for col_index, val in enumerate(normalized, start=1):
            cell = sheet.cell(row=append_row_idx, column=col_index)
            rule = self.validation_rules[col_index - 1] if col_index - 1 < len(self.validation_rules) else {"type": "text"}
            if val is None:
                cell.value = None
            elif rule.get("type") == "numeric" and isinstance(val, (int, float)):
                cell.value = val
            elif rule.get("type") == "date":
                try:
                    cell.value = datetime.fromisoformat(val).date()
                except Exception:
                    cell.value = val
            else:
                cell.value = val

        # Reflect in UI
        display_row = [(v if v is not None else "") for v in normalized]
        self.tree.insert("", tk.END, values=display_row)

        self.unsaved_changes = True
        self._update_status(f"Added new row to '{self.active_sheet_name}'.")

        # Auto-save if enabled
        if self.auto_save_var.get():
            if self.save_file():
                self._update_status(f"Auto-saved after adding new row ✅")

        # Reset form
        self.clear_input_entries()
        if self.input_entries:
            self.input_entries[0].focus_set()

    # -----------------------
    # Editable Treeview Rows
    # -----------------------
    def on_tree_double_click(self, event):
        """Enable editing mode by loading selected row data into input fields."""
        selected_item = self.tree.focus()
        if not selected_item:
            return

        values = self.tree.item(selected_item, "values")
        if not values:
            return

        # Populate inputs
        for entry, value in zip(self.input_entries, values):
            entry.delete(0, tk.END)
            entry.insert(0, value)

        self.editing_item = selected_item
        self.mode = "edit"

        # Change button text + command
        self.add_button.config(text="Update Row", command=self.update_row_from_inputs)

        # ✅ Temporarily unbind <Return> from adding rows
        for entry in self.input_entries:
            entry.unbind("<Return>")
            entry.bind("<Return>", lambda e: self.update_row_from_inputs())

        # Focus first field
        if self.input_entries:
            self.input_entries[0].focus_set()

        self._update_status("Editing existing row...")

    def update_row_from_inputs(self):
        """Update selected Treeview row and workbook entry."""
        if not hasattr(self, "editing_item") or not self.editing_item:
            messagebox.showinfo("No selection", "No row selected for editing.")
            return

        is_valid, messages, normalized = self.validate_inputs()
        if not is_valid:
            messagebox.showwarning("Validation failed", "\n".join(messages))
            return

        # Update Treeview
        self.tree.item(self.editing_item, values=[(v if v is not None else "") for v in normalized])

        # Update workbook
        sheet = self.workbook[self.active_sheet_name]
        row_index = self.tree.index(self.editing_item) + 2  # +2 for header row
        for col_index, val in enumerate(normalized, start=1):
            rule = self.validation_rules[col_index - 1] if col_index - 1 < len(self.validation_rules) else {"type": "text"}
            cell = sheet.cell(row=row_index, column=col_index)
            if val is None:
                cell.value = None
            elif rule.get("type") == "numeric" and isinstance(val, (int, float)):
                cell.value = val
            elif rule.get("type") == "date":
                try:
                    cell.value = datetime.fromisoformat(val).date()
                except Exception:
                    cell.value = val
            else:
                cell.value = val

        self.unsaved_changes = True
        self._update_status(f"Updated row {row_index - 1} successfully.")

        # ✅ Re-highlight the updated row
        self.tree.selection_set(self.editing_item)
        self.tree.see(self.editing_item)

        # ✅ Auto-save if enabled
        if self.auto_save_var.get():
            if self.save_file():
                self._update_status("Auto-saved after editing row ✅")

        # ✅ Reset UI and rebind Enter key for Add mode
        self.clear_input_entries()
        self.add_button.config(text="Add Row", command=self.add_row_from_inputs)
        self.mode = "add"
        self.editing_item = None

        for entry in self.input_entries:
            entry.unbind("<Return>")
            entry.bind("<Return>", lambda e: self.add_row_from_inputs())

        if self.input_entries:
            self.input_entries[0].focus_set()

    # -----------------------
    # Save Operations
    # -----------------------
    def save_file(self):
        if not self.workbook:
            messagebox.showwarning("No file", "No workbook is currently open.")
            return False

        if not self.filepath:
            return self.save_file_as()

        try:
            self.workbook.save(self.filepath)
            self.unsaved_changes = False
            self._update_status(f"Saved: {os.path.basename(self.filepath)}")
            return True
        except Exception as e:
            messagebox.showerror("Save error", f"Failed to save workbook:\n{e}")
            return False

    def save_file_as(self):
        filetypes = [("Excel files", "*.xlsx")]
        path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=filetypes)
        if not path:
            return False
        self.filepath = path
        try:
            self.workbook.save(self.filepath)
            self.unsaved_changes = False
            self._update_status(f"Saved as: {os.path.basename(self.filepath)}")
            return True
        except Exception as e:
            messagebox.showerror("Save error", f"Failed to save workbook:\n{e}")
            return False

    # -----------------------
    # Theme handling
    # -----------------------
    def on_theme_change(self, event=None):
        theme_name = self.theme_combo.get()
        try:
            Style().theme_use(theme_name)
            self._update_status(f"Theme changed to {theme_name}")
        except Exception as e:
            messagebox.showerror("Theme error", f"Cannot set theme {theme_name}:\n{e}")

    # -----------------------
    # Helpers & closing
    # -----------------------
    def _update_status(self, text):
        self.status_var.set(text)

    def _show_about(self):
        messagebox.showinfo("About", f"{APP_TITLE}\nDynamic Excel-driven data entry with adaptive validation.\nBuilt with ttkbootstrap + openpyxl")

    def on_close(self):
        if self.unsaved_changes:
            res = messagebox.askyesnocancel("Unsaved changes", "You have unsaved changes. Save before exit?")
            if res is None:
                return
            if res:
                if not self.save_file():
                    return

        try:
            self.root.destroy()
        except Exception:
            os._exit(0)


def main():
    app_root = Window(title=APP_TITLE, themename="cosmo")
    app = DynamicExcelApp(app_root)

    # --- Add Auto-Save Checkbox ---
    app.auto_save_var = tk.BooleanVar(value=False)
    auto_save_chk = ttk.Checkbutton(app_root, text="Auto-Save on Add/Edit", variable=app.auto_save_var)
    auto_save_chk.pack(pady=3)

    # Bind double-click editing
    app.tree.bind("<Double-1>", app.on_tree_double_click)

    app_root.mainloop()


if __name__ == "__main__":
    main()


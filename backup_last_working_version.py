# data_entry_v2_dynamic_excel_with_validation.py
# v2+: Dynamic Excel-driven Tkinter data-entry app with adaptive validation
#
# Save as: data_entry_v2_dynamic_excel_with_validation.py
# Requirements:
# pip install ttkbootstrap openpyxl

import os
import re
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from ttkbootstrap import Window, Style
from openpyxl import load_workbook, Workbook
from datetime import datetime, date

APP_TITLE = "Data Entry v2+ — Dynamic Excel Companion (with Validation)"


# -------------------------
# Validation helpers
# -------------------------
EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")

def try_parse_date(value: str):
    """Try multiple common date formats. Return date object if successful, else None."""
    if value is None:
        return None
    s = value.strip()
    if s == "":
        return None
    formats = ["%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y", "%Y/%m/%d"]
    for fmt in formats:
        try:
            return datetime.strptime(s, fmt).date()
        except Exception:
            continue
    # try ISO fallback
    try:
        return datetime.fromisoformat(s).date()
    except Exception:
        return None

def is_numeric(value: str):
    if value is None:
        return False
    s = str(value).strip()
    if s == "":
        return False
    try:
        float(s.replace(",", ""))  # allow commas in thousands
        return True
    except Exception:
        return False

def normalize_numeric(value: str):
    return float(str(value).strip().replace(",", ""))


# -------------------------
# Main app
# -------------------------
class DynamicExcelApp:
    def __init__(self, root: Window):
        self.root = root
        self.root.title(APP_TITLE)
        self.root.geometry("1100x720")

        # State
        self.workbook = None
        self.filepath = None
        self.active_sheet_name = None
        self.headers = []
        self.input_entries = []  # list of tk.Entry widgets matching headers
        self.unsaved_changes = False
        self.original_editing_values = {} # NEW: Store original values for uniqueness check exclusion
        self.mode = "add" # Can be "add" or "edit"

        # Inferred validation rules per column: list of dicts
        self.validation_rules = []

        # UI elements
        self._create_menu()
        self._create_toolbar()
        self._create_top_frame()
        self._create_bottom_frame()
        self._create_statusbar()
        self._bind_events()
        
        # NOTE: Auto-save var is attached in the `main` function below the class definition

        # Start with a clean app — prompt user to open file
        self._prompt_open_file_on_startup()

    # -----------------------
    # UI Construction
    # -----------------------
    def _create_menu(self):
        menubar = tk.Menu(self.root)
        file_menu = tk.Menu(menubar, tearoff=0)
        file_menu.add_command(label="Open...", command=self.open_file)
        file_menu.add_command(label="Save", command=self.save_file)
        file_menu.add_command(label="Save As...", command=self.save_file_as)
        file_menu.add_separator()
        file_menu.add_command(label="Exit", command=self.on_close)
        menubar.add_cascade(label="File", menu=file_menu)

        help_menu = tk.Menu(menubar, tearoff=0)
        help_menu.add_command(label="About", command=self._show_about)
        menubar.add_cascade(label="Help", menu=help_menu)

        self.root.config(menu=menubar)

    def _create_toolbar(self):
        toolbar = ttk.Frame(self.root)
        toolbar.pack(side=tk.TOP, fill=tk.X, padx=6, pady=6)

        open_btn = ttk.Button(toolbar, text="Open File", command=self.open_file)
        open_btn.pack(side=tk.LEFT, padx=(0, 6))

        save_btn = ttk.Button(toolbar, text="Save", command=self.save_file)
        save_btn.pack(side=tk.LEFT, padx=(0, 6))

        # --- New: Delete Row Button ---
        delete_btn = ttk.Button(toolbar, text="Delete Selected Row", command=self.delete_selected_row, style="danger.TButton")
        delete_btn.pack(side=tk.LEFT, padx=(0, 12))

        self.add_button = ttk.Button(toolbar, text="Add Row", command=self.add_row_from_inputs, style="success.TButton")
        self.add_button.pack(side=tk.LEFT, padx=(0, 12))

        ttk.Label(toolbar, text="Sheet:").pack(side=tk.LEFT, padx=(12, 4))
        self.sheet_combo = ttk.Combobox(toolbar, state="readonly", width=30)
        self.sheet_combo.pack(side=tk.LEFT)
        self.sheet_combo.bind("<<ComboboxSelected>>", self.on_sheet_change)

        # Spacer (removed status label from here)
        spacer = ttk.Label(toolbar, text="")
        spacer.pack(side=tk.LEFT, expand=True)

        # Theme toggler
        ttk.Label(toolbar, text="Theme:").pack(side=tk.LEFT, padx=(6, 4))
        self.theme_combo = ttk.Combobox(toolbar, values=Style().theme_names(), state="readonly", width=15)
        self.theme_combo.set(Style().theme_use())
        self.theme_combo.bind("<<ComboboxSelected>>", self.on_theme_change)
        self.theme_combo.pack(side=tk.LEFT, padx=(0, 6))

    def _create_statusbar(self):
        # The status bar replaces the old status label
        self.status_var = tk.StringVar(value="No file opened.")
        statusbar = ttk.Label(self.root, textvariable=self.status_var, relief=tk.SUNKEN, anchor=tk.W)
        statusbar.pack(side=tk.BOTTOM, fill=tk.X)

    def _create_top_frame(self):
        # top frame will contain dynamic input fields (headers -> entries)
        self.top_frame = ttk.Frame(self.root)
        self.top_frame.pack(side=tk.TOP, fill=tk.X, padx=10, pady=(6, 0))

        # We'll put inputs into a scrollable frame in case there are many columns
        self.input_canvas = tk.Canvas(self.top_frame, height=140)
        self.input_canvas.pack(side=tk.TOP, fill=tk.X, expand=True)

        self.input_scrollbar = ttk.Scrollbar(self.top_frame, orient=tk.HORIZONTAL, command=self.input_canvas.xview)
        self.input_scrollbar.pack(side=tk.TOP, fill=tk.X)
        self.input_canvas.configure(xscrollcommand=self.input_scrollbar.set)

        self.inputs_inner = ttk.Frame(self.input_canvas)
        self.input_canvas.create_window((0, 0), window=self.inputs_inner, anchor="nw")
        self.inputs_inner.bind("<Configure>", lambda e: self.input_canvas.configure(scrollregion=self.input_canvas.bbox("all")))

    def _create_bottom_frame(self):
        # bottom frame will contain the treeview of sheet data
        bottom_frame = ttk.Frame(self.root)
        bottom_frame.pack(side=tk.TOP, fill=tk.BOTH, expand=True, padx=10, pady=10)

        self.tree = ttk.Treeview(bottom_frame, show="headings")
        self.tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        vsb = ttk.Scrollbar(bottom_frame, orient="vertical", command=self.tree.yview)
        vsb.pack(side=tk.LEFT, fill=tk.Y)
        self.tree.configure(yscrollcommand=vsb.set)

        hsb = ttk.Scrollbar(bottom_frame, orient="horizontal", command=self.tree.xview)
        hsb.pack(side=tk.BOTTOM, fill=tk.X)
        self.tree.configure(xscrollcommand=hsb.set)

    def _bind_events(self):
        # Save prompt on close
        self.root.protocol("WM_DELETE_WINDOW", self.on_close)
        # Bind double-click editing
        self.tree.bind("<Double-1>", self.on_tree_double_click)


    # -----------------------
    # File and Sheet Handling
    # -----------------------
    def _prompt_open_file_on_startup(self):
        # Clean workspace: prompt for file
        answer = messagebox.askyesno("Open file", "Do you want to open an existing .xlsx file to work on?")
        if answer:
            self.open_file()
        else:
            self.status_var.set("Ready. Use File -> Open to open an .xlsx file.")

    def open_file(self):
        filetypes = [("Excel files", "*.xlsx")]
        path = filedialog.askopenfilename(title="Open Excel file", filetypes=filetypes)
        if not path:
            return

        try:
            wb = load_workbook(path)
        except Exception as e:
            messagebox.showerror("Error", f"Failed to open workbook:\n{e}")
            return

        self.workbook = wb
        self.filepath = path
        # default to active sheet
        self.active_sheet_name = self.workbook.active.title
        self._populate_sheet_selector()
        self._load_active_sheet()
        self.unsaved_changes = False
        self._update_status(f"Opened: {os.path.basename(path)}")

    def _populate_sheet_selector(self):
        if not self.workbook:
            self.sheet_combo["values"] = []
            return
        names = self.workbook.sheetnames
        self.sheet_combo["values"] = names
        if self.active_sheet_name in names:
            self.sheet_combo.set(self.active_sheet_name)
        else:
            self.sheet_combo.set(names[0])
            self.active_sheet_name = names[0]

    def on_sheet_change(self, event=None):
        if not self.workbook:
            return
        new_sheet = self.sheet_combo.get()
        if new_sheet == self.active_sheet_name:
            return

        if self.unsaved_changes:
            res = messagebox.askyesnocancel("Unsaved changes", "You have unsaved changes. Save before switching sheets?")
            if res is None:
                # Cancel switching
                self.sheet_combo.set(self.active_sheet_name)
                return
            if res:  # Yes -> save
                if not self.save_file():
                    # Save failed or cancelled; cancel switching
                    self.sheet_combo.set(self.active_sheet_name)
                    return

        self.active_sheet_name = new_sheet
        self._load_active_sheet()

    def _load_active_sheet(self):
        """Read headers and rows from active sheet, rebuild inputs and treeview."""
        if not self.workbook or not self.active_sheet_name:
            return

        sheet = self.workbook[self.active_sheet_name]
        # Read headers: find first non-empty row (usually row 1). We treat the very first row with any non-empty cells as headers.
        headers = []
        header_row_idx = None
        for r in sheet.iter_rows(min_row=1, max_row=5):  # check first few rows for header row
            values = [cell.value for cell in r]
            if any(v is not None and str(v).strip() != "" for v in values):
                header_row_idx = r[0].row
                headers = []
                for cell in r:
                    # Stop at trailing blank cells? We'll keep blanks but give them fallback names
                    val = cell.value
                    headers.append(str(val).strip() if val is not None and str(val).strip() != "" else None)
                break

        if not headers:
            # If still no headers, create default Column1..N based on max_column
            max_col = sheet.max_column or 1
            headers = [None] * max_col
            header_row_idx = 1

        # Normalize headers: fill None with "Column X"
        normalized = []
        for i, h in enumerate(headers):
            if h:
                normalized.append(h)
            else:
                normalized.append(f"Column {i+1}")
        self.headers = normalized

        # Infer validation rules
        self.validation_rules = self._infer_validation_rules(self.headers)

        # Build inputs and treeview
        self._build_input_fields(self.headers)
        self._load_treeview_rows(sheet, header_row_idx)

    # -----------------------
    # Validation rule inference (UPDATED)
    # -----------------------
    # In DynamicExcelApp._infer_validation_rules(self, headers)

    def _infer_validation_rules(self, headers):
        """
        Heuristic inference of validation rules from header names, including uniqueness.
        Returns a list of dicts: {'name': str, 'type': 'text'|..., 'required_var': tk.BVar, 'duplicate_var': tk.SVar, ...} 
        """
        rules = []
        numeric_keywords = ("qty", "quantity", "amount", "price", "total", "number", "age", "count")
        
        for h in headers:
            h_lower = h.lower() if h else ""
            
            # --- Initialize default rule state ---
            is_required_default = True
            duplicate_policy_default = "none"
            val_type_default = "text"
            
            # 1. Determine Default Required Status
            if "optional" in h_lower or "[optional]" in h_lower or "(optional)" in h_lower:
                is_required_default = False
            
            # 2. Determine Default Duplicate Policy
            if "(unique)" in h_lower or "[strict]" in h_lower or "id" in h_lower: 
                duplicate_policy_default = "strict"
            elif "(duplicate-warn)" in h_lower or "[warn]" in h_lower:
                duplicate_policy_default = "warn"

            # 3. Determine Default Type
            if any(keyword in h_lower for keyword in numeric_keywords):
                val_type_default = "numeric"
            elif "date" in h_lower:
                val_type_default = "date"
            elif "email" in h_lower:
                val_type_default = "email"
                
            # Final check for text fields often optional
            if val_type_default == "text" and ("description" in h_lower or "notes" in h_lower):
                 is_required_default = False
            
            # If it's a unique ID, we generally make it optional if it might be auto-generated
            if val_type_default == "text" and "id" in h_lower and duplicate_policy_default == "strict": 
                 is_required_default = False
                 
            # --- Create TK Variables and Final Rule Dict ---
            
            # We use is_required_default for initialization, 
            # and 'required' (in the dict) for the current state.
            rule = {
                "name": h, 
                "type": val_type_default, 
                # These vars hold the user's current selection
                "required_var": tk.BooleanVar(value=is_required_default), 
                "duplicate_var": tk.StringVar(value=duplicate_policy_default),
                # This stores the inferred default state to use in validation
                "required": is_required_default,
                "duplicate_policy": duplicate_policy_default
            }
            
            rules.append(rule)
        return rules

    # -----------------------
    # Inputs builder & validation (UPDATED)
    # -----------------------
    def _clear_inputs_area(self):
        for w in self.inputs_inner.winfo_children():
            w.destroy()
        self.input_entries.clear()

    def _build_input_fields(self, headers):
        """Create a horizontal layout of label+entry, plus QoL controls for each header."""
        self._clear_inputs_area()

        for idx, header in enumerate(headers):
            rule = self.validation_rules[idx]
            
            col_frame = ttk.Frame(self.inputs_inner)
            col_frame.grid(row=0, column=idx, padx=6, pady=4)
            
            # --- Label improvement: Show current state (R/U/W) ---
            # This is complex to do dynamically, so we'll simplify the label and rely on the controls below.
            lbl = ttk.Label(col_frame, text=header, width=20, anchor="center")
            lbl.pack(side=tk.TOP, fill=tk.X)
            
            # Entry and Error Label (existing logic)
            ent = tk.Entry(col_frame, width=20)
            ent.pack(side=tk.TOP, pady=(6, 0))
            ent.bind("<Return>", lambda e, i=idx: self._on_enter_pressed(e, i)) 
            self.input_entries.append(ent)
            
            error_var = tk.StringVar(value="")
            error_lbl = ttk.Label(col_frame, textvariable=error_var, foreground="red", anchor="center")
            error_lbl.pack(side=tk.TOP, fill=tk.X)
            ent.error_var = error_var 
            
            # --- NEW QoL Controls Frame ---
            control_frame = ttk.Frame(col_frame)
            control_frame.pack(side=tk.TOP, fill=tk.X, pady=(5, 0))
            
            # 1. Required Checkbox
            req_chk = ttk.Checkbutton(control_frame, text="Required", variable=rule['required_var'], 
                                      command=lambda r=rule: self._update_validation_state(r))
            req_chk.pack(anchor=tk.W)

            # 2. Duplicate Radio Buttons
            dup_lbl = ttk.Label(control_frame, text="Duplicate Policy:", style="TLabel")
            dup_lbl.pack(anchor=tk.W, pady=(2, 0))
            
            # Radio button for None
            ttk.Radiobutton(control_frame, text="None", variable=rule['duplicate_var'], value="none",
                            command=lambda r=rule: self._update_validation_state(r)).pack(anchor=tk.W, padx=10)
            
            # Radio button for Warn
            ttk.Radiobutton(control_frame, text="Warn", variable=rule['duplicate_var'], value="warn",
                            command=lambda r=rule: self._update_validation_state(r)).pack(anchor=tk.W, padx=10)
                            
            # Radio button for Strict
            ttk.Radiobutton(control_frame, text="Strict", variable=rule['duplicate_var'], value="strict",
                            command=lambda r=rule: self._update_validation_state(r)).pack(anchor=tk.W, padx=10)
            
        self.root.after(100, lambda: self.input_canvas.configure(scrollregion=self.input_canvas.bbox("all")))
        self.reset_to_add_mode()

    def _update_validation_state(self, rule):
        """Updates the internal validation rule dictionary from the user's QoL controls."""
        
        # Read the current Tk variable states
        new_required = rule['required_var'].get()
        new_duplicate_policy = rule['duplicate_var'].get()
        
        # Update the rule dictionary used by validate_inputs
        rule["required"] = new_required
        rule["duplicate_policy"] = new_duplicate_policy
        
        self._update_status(f"Validation policy updated for '{rule['name']}'. Required: {new_required}, Duplicate: {new_duplicate_policy}")

    def _on_enter_pressed(self, event, idx):
        # If Enter pressed in last field, add row; else focus next.
        if idx == len(self.input_entries) - 1:
            if self.mode == "add":
                self.add_row_from_inputs()
            elif self.mode == "edit":
                 self.update_row_from_inputs()
        else:
            self.input_entries[idx + 1].focus_set()


    def _get_existing_column_data(self, col_index):
        """
        NEW HELPER: Retrieves all non-empty, normalized values from a specific column 
        in the active sheet (data rows only). Returns a set for fast lookup.
        """
        if not self.workbook or not self.active_sheet_name:
            return set()
            
        sheet = self.workbook[self.active_sheet_name]
        
        # We assume the data starts at row 2 (after the header row)
        existing_values = set()
        
        # Iterate over all rows starting from the data rows (row 2 or higher)
        for row_idx in range(2, sheet.max_row + 1): 
            cell_value = sheet.cell(row=row_idx, column=col_index).value
            if cell_value is not None:
                # Normalize the value (strip whitespace, convert to string) for case-insensitive comparison
                normalized_value = str(cell_value).strip()
                if normalized_value:
                    existing_values.add(normalized_value.lower()) 
                    
        return existing_values
        
    def validate_inputs(self):
        """
        Validate all input entries. Applies background color and error labels for feedback.
        Returns (is_valid: bool, strict_messages: list, warning_messages: list, normalized_values: list)
        """
        normalized = []
        strict_messages = []
        warning_messages = [] 
        is_valid = True
        
        is_edit_mode = self.mode == "edit"

        for i, entry in enumerate(self.input_entries):
            val = entry.get()
            rule = self.validation_rules[i]
            col_name = rule["name"]
            val_type = rule["type"]
            # --- CRITICAL: USE UPDATED STATE KEYS ---
            required = rule["required"] 
            duplicate_policy = rule["duplicate_policy"]
            
            # Reset feedback for this entry
            entry.config(bg="white")
            if hasattr(entry, 'error_var'):
                entry.error_var.set("")
            
            val_stripped = val.strip()
            
            # 1. Required check
            if required and not val_stripped:
                is_valid = False
                strict_messages.append(f"❌ {col_name}: Required.") # <-- CORRECTED
                normalized.append(None)
                entry.config(bg="#fbb") # Light red background for error
                if hasattr(entry, 'error_var'):
                    entry.error_var.set("Required")
                continue # Move to next field

            # Skip type validation for optional, empty fields
            if not val_stripped and not required:
                 normalized.append(None)
                 continue
            
            # 2. Uniqueness Check (MODIFIED LOGIC)
            if duplicate_policy in ("strict", "warn"):
                col_index = i + 1
                existing_data = self._get_existing_column_data(col_index=col_index)
                
                # Exclusion Logic for Edit Mode
                is_original_value = False
                if is_edit_mode:
                    original_val = self.original_editing_values.get(col_index, "__NO_MATCH__")
                    # Check against the normalized original value
                    if val_stripped.lower() == original_val:
                        is_original_value = True
                        
                # Only check for duplication if it's NOT the original value
                if not is_original_value and val_stripped.lower() in existing_data:
                    # DUPLICATE VIOLATION FOUND
                    if duplicate_policy == "strict":
                        # Treat as an error
                        is_valid = False
                        strict_messages.append(f"❌ {col_name}: Duplicate value found (Strict Policy).")
                        entry.config(bg="#fbb")
                        if hasattr(entry, 'error_var'):
                            entry.error_var.set("Duplicate (Strict)")
                        normalized.append(val_stripped)
                        continue # Stop further checks for this field
                    
                    elif duplicate_policy == "warn":
                        # Treat as a warning, but keep is_valid = True
                        warning_messages.append(f"⚠️ {col_name}: Possible duplicate value found.")
                        entry.config(bg="#ffdd99") # Yellow/orange background for warning
                        if hasattr(entry, 'error_var'):
                            entry.error_var.set("Possible Duplicate")
                        # DO NOT continue, proceed to type check
            
            # 3. Type-specific validation and normalization
            try:
                if val_type == "text":
                    cleaned = re.sub(r"\s+", " ", val_stripped)
                    normalized.append(cleaned if cleaned else None)
                    
                elif val_type == "numeric":
                    if not is_numeric(val):
                        raise ValueError("Invalid number")
                    normalized.append(normalize_numeric(val)) 

                elif val_type == "date":
                    date_obj = try_parse_date(val)
                    if date_obj is None:
                        raise ValueError("Invalid date (Try YYYY-MM-DD)")
                    normalized.append(date_obj) 

                elif val_type == "email":
                    if not EMAIL_RE.match(val_stripped):
                        raise ValueError("Invalid email format")
                    normalized.append(val_stripped)

                else:
                    normalized.append(val_stripped)

            except ValueError as e:
                is_valid = False
                strict_messages.append(f"❌ {col_name}: {e}") # <-- CORRECTED
                entry.config(bg="#fbb")
                if hasattr(entry, 'error_var'):
                    entry.error_var.set(str(e))
                normalized.append(val_stripped)

        return is_valid, strict_messages, warning_messages, normalized

    def clear_input_entries(self):
        for ent in self.input_entries:
            ent.delete(0, tk.END)
            # Reset background and error label
            ent.config(bg="white")
            if hasattr(ent, 'error_var'):
                ent.error_var.set("")

    # -----------------------
    # Treeview population
    # -----------------------
    def _clear_treeview(self):
        for col in self.tree["columns"]:
            self.tree.heading(col, text="")
        self.tree.delete(*self.tree.get_children())
        self.tree["columns"] = ()

    def _load_treeview_rows(self, sheet, header_row_idx):
        self._clear_treeview()
        # Treeview columns -> headers
        cols = [f"c{i}" for i in range(len(self.headers))]
        self.tree["columns"] = cols
        for i, h in enumerate(self.headers):
            self.tree.heading(cols[i], text=h, anchor=tk.W)
            # set column width heuristically
            self.tree.column(cols[i], width=160, anchor=tk.W)

        # Read rows starting after header_row_idx
        start_row = header_row_idx + 1
        rows = []
        for r in sheet.iter_rows(min_row=start_row, max_row=sheet.max_row, max_col=len(self.headers)):
            rowvals = [ (cell.value if cell.value is not None else "") for cell in r ]
            # If the row is entirely empty, skip
            if all((v == "" or v is None) for v in rowvals):
                continue
            rows.append(rowvals)

        for row in rows:
            # Ensure row length matches headers
            row_extended = list(row) + [""] * (len(self.headers) - len(row))
            self.tree.insert("", tk.END, values=row_extended)

    # -----------------------
    # Add / Edit / Delete Row
    # -----------------------
    def add_row_from_inputs(self):
        """Append a new row after validating input fields."""
        if not self.workbook or not self.active_sheet_name:
            messagebox.showwarning("No file", "Open an .xlsx file first.")
            return

        # Validate inputs # --- CRITICAL CHANGE: Capture warnings list ---
        is_valid, strict_messages, warning_messages, normalized = self.validate_inputs()

        # 1. Handle Strict Validation Failure
        if not is_valid:
            self._update_status(f"Strict validation failed on {len(strict_messages)} field(s).")
            return
            
        # 2. Handle Duplicate Warnings and Prompt User
        if warning_messages:
            warning_text = "\n".join(warning_messages)
            prompt = (
                "The following potential duplicate entries were detected:\n\n"
                f"{warning_text}\n\n"
                "Do you still want to add this record?"
            )
            res = messagebox.askyesno("Possible Duplicate Detected", prompt, icon='warning')
            
            if not res:
                self._update_status("Addition cancelled due to duplicate warning.")
                return # User chose not to proceed

        sheet = self.workbook[self.active_sheet_name]
        append_row_idx = sheet.max_row + 1  
        
        # --- Write values into workbook ---
        for col_index, val in enumerate(normalized, start=1):
            cell = sheet.cell(row=append_row_idx, column=col_index)
            cell.value = val 

        # Reflect in UI: Treeview needs string representation for display
        display_row = []
        for val in normalized:
            if val is None:
                display_row.append("")
            elif isinstance(val, date):
                # Use strftime to format the date as YYYY-MM-DD
                display_row.append(val.strftime("%Y-%m-%d"))
            elif isinstance(val, (int, float)):
                 display_row.append(str(val)) # Convert number to string
            else:
                display_row.append(str(val))

        # Highlight newly-added row
        new_item_id = self.tree.insert("", tk.END, values=display_row)
        self.tree.selection_remove(self.tree.selection())  
        self.tree.selection_set(new_item_id)
        self.tree.see(new_item_id)

        self.unsaved_changes = True
        self._update_status(f"Added new row to '{self.active_sheet_name}'.")

        # Auto-save if enabled
        if hasattr(self, 'auto_save_var') and self.auto_save_var.get():
            if self.save_file():
                self._update_status(f"Auto-saved after adding new row ✅")

        # Reset form
        self.clear_input_entries()
        if self.input_entries:
            self.input_entries[0].focus_set()

    def on_tree_double_click(self, event):
        """Enable editing mode by loading selected row data into input fields."""
        selected_item = self.tree.focus()
        if not selected_item:
            return

        values = self.tree.item(selected_item, "values")
        if not values:
            return

        # Clear previous error feedback before populating new data
        self.clear_input_entries()

        # Populate inputs
        for entry, value in zip(self.input_entries, values):
            entry.delete(0, tk.END)
            entry.insert(0, value)

        self.editing_item = selected_item
        self.mode = "edit"

        # --- NEW: Store original values for uniqueness check exclusion ---
        self.original_editing_values = {}
        for i, val in enumerate(values):
            # Store the normalized (lowercase, stripped) value of the original cell data
            self.original_editing_values[i + 1] = str(val).strip().lower() 
            
        # Change button text + command
        self.add_button.config(text="Update Row", command=self.update_row_from_inputs, style="warning.TButton")

        # Temporarily unbind <Return> from its original 'focus next' logic
        for entry in self.input_entries:
            entry.unbind("<Return>")
            # Bind to a simple update call
            entry.bind("<Return>", lambda e: self.update_row_from_inputs())

        # Focus first field
        if self.input_entries:
            self.input_entries[0].focus_set()

        self._update_status("Editing existing row...")

    def update_row_from_inputs(self):
        """Update selected Treeview row and workbook entry."""
        if not hasattr(self, "editing_item") or not self.editing_item:
            messagebox.showinfo("No selection", "No row selected for editing.")
            return

        # --- CRITICAL CHANGE: Capture warnings list ---
        is_valid, strict_messages, warning_messages, normalized = self.validate_inputs()
        
        # 1. Handle Strict Validation Failure
        if not is_valid:
            self._update_status(f"Strict validation failed on {len(strict_messages)} fields. See red fields.")
            return

        # 2. Handle Duplicate Warnings and Prompt User
        if warning_messages:
            warning_text = "\n".join(warning_messages)
            prompt = (
                "The following potential duplicate entries were detected:\n\n"
                f"{warning_text}\n\n"
                "Do you still want to update this record?"
            )
            res = messagebox.askyesno("Possible Duplicate Detected", prompt, icon='warning')
            
            if not res:
                self._update_status("Update cancelled due to duplicate warning.")
                return # User chose not to proceed

        # Update Treeview - same conversion to display strings as in add_row
        display_row = []
        for val in normalized:
            if val is None:
                display_row.append("")
            elif isinstance(val, date):
                # Use strftime to format the date as YYYY-MM-DD, removing any time component
                display_row.append(val.strftime("%Y-%m-%d"))
            elif isinstance(val, (int, float)):
                 display_row.append(str(val)) 
            else:
                display_row.append(str(val))

        self.tree.item(self.editing_item, values=display_row) 

        # Update workbook - use normalized values (ready for openpyxl)
        sheet = self.workbook[self.active_sheet_name]
        # +2 for header row (assuming header is in row 1)
        row_index = self.tree.index(self.editing_item) + 2 
        
        for col_index, val in enumerate(normalized, start=1):
            cell = sheet.cell(row=row_index, column=col_index)
            cell.value = val 

        self.unsaved_changes = True
        self._update_status(f"Updated row {row_index - 1} successfully.")

        # Re-highlight the updated row
        self.tree.selection_set(self.editing_item)
        self.tree.see(self.editing_item)

        # Auto-save if enabled
        if hasattr(self, 'auto_save_var') and self.auto_save_var.get():
            if self.save_file():
                self._update_status("Auto-saved after editing row ✅")

        # Reset UI and rebind Enter key for Add mode
        self.reset_to_add_mode()

    def delete_selected_row(self):
        """Deletes the selected row from the Treeview and the in-memory Excel sheet."""
        if not self.workbook or not self.active_sheet_name:
            messagebox.showwarning("No file", "Open an .xlsx file first.")
            return

        selected_item = self.tree.focus()
        if not selected_item:
            messagebox.showinfo("No selection", "Please select a row to delete.")
            return

        # Confirmation dialog
        confirm = messagebox.askyesno("Confirm Deletion", "Are you sure you want to delete the selected row? This action is irreversible before saving.")
        if not confirm:
            return

        try:
            # 1. Determine Excel row index
            sheet = self.workbook[self.active_sheet_name]
            excel_row_index = self.tree.index(selected_item) + 2  # +1 for header +1 for index offset

            # 2. Remove row from Excel (openpyxl)
            sheet.delete_rows(excel_row_index, 1)

            # 3. Remove row from Treeview (Tkinter)
            self.tree.delete(selected_item)

            self.unsaved_changes = True
            self._update_status(f"Deleted row from '{self.active_sheet_name}'.")

            # Reset to add mode if we were editing the deleted row
            if hasattr(self, "editing_item") and self.editing_item == selected_item:
                self.reset_to_add_mode()
                
            # Auto-save if enabled
            if hasattr(self, 'auto_save_var') and self.auto_save_var.get():
                if self.save_file():
                    self._update_status("Auto-saved after deleting row ✅")

        except Exception as e:
            messagebox.showerror("Error", f"Failed to delete row:\n{e}")

    def reset_to_add_mode(self):
        """Helper to switch back to the default 'Add Row' state."""
        self.clear_input_entries()
        self.add_button.config(text="Add Row", command=self.add_row_from_inputs, style="success.TButton")
        
        # Re-bind Enter key for Add mode
        for idx, entry in enumerate(self.input_entries):
             entry.unbind("<Return>")
             # Rebind the original logic (focus next or add row)
             entry.bind("<Return>", lambda e, i=idx: self._on_enter_pressed(e, i)) 
        
        self.mode = "add"
        self.editing_item = None
        self.original_editing_values = {} # Clear stored original values
        if self.input_entries:
            self.input_entries[0].focus_set()

    # -----------------------
    # Save Operations
    # -----------------------
    def save_file(self):
        if not self.workbook:
            messagebox.showwarning("No file", "No workbook is currently open.")
            return False

        if not self.filepath:
            return self.save_file_as()

        try:
            self.workbook.save(self.filepath)
            self.unsaved_changes = False
            self._update_status(f"Saved: {os.path.basename(self.filepath)}")
            return True
        except Exception as e:
            messagebox.showerror("Save error", f"Failed to save workbook:\n{e}")
            return False

    def save_file_as(self):
        filetypes = [("Excel files", "*.xlsx")]
        path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=filetypes)
        if not path:
            return False
        self.filepath = path
        try:
            self.workbook.save(self.filepath)
            self.unsaved_changes = False
            self._update_status(f"Saved as: {os.path.basename(self.filepath)}")
            return True
        except Exception as e:
            messagebox.showerror("Save error", f"Failed to save workbook:\n{e}")
            return False

    # -----------------------
    # Theme handling
    # -----------------------
    def on_theme_change(self, event=None):
        theme_name = self.theme_combo.get()
        try:
            Style().theme_use(theme_name)
            self._update_status(f"Theme changed to {theme_name}")
        except Exception as e:
            messagebox.showerror("Theme error", f"Cannot set theme {theme_name}:\n{e}")

    # -----------------------
    # Helpers & closing
    # -----------------------
    def _update_status(self, text):
        self.status_var.set(text)

    def _show_about(self):
        messagebox.showinfo("About", f"{APP_TITLE}\nDynamic Excel-driven data entry with adaptive validation.\nBuilt with ttkbootstrap + openpyxl")

    def on_close(self):
        if self.unsaved_changes:
            res = messagebox.askyesnocancel("Unsaved changes", "You have unsaved changes. Save before exit?")
            if res is None:
                return
            if res:
                if not self.save_file():
                    return

        try:
            self.root.destroy()
        except Exception:
            os._exit(0)
            
# -------------------------
# Application Entry Point
# -------------------------
def main():
    app_root = Window(title=APP_TITLE, themename="cosmo")
    app = DynamicExcelApp(app_root)

    # --- Add Auto-Save Checkbox ---
    # Attach to the app object for easy access in its methods
    app.auto_save_var = tk.BooleanVar(value=False) 
    auto_save_chk = ttk.Checkbutton(app_root, text="Auto-Save on Add/Edit/Delete", 
                                    variable=app.auto_save_var, style="primary.TCheckbutton")
    auto_save_chk.pack(side=tk.TOP, anchor=tk.NW, padx=10, pady=(0, 6)) # Positioned below toolbar

    app_root.mainloop()


if __name__ == "__main__":
    main()
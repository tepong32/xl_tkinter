# data_entry_v2_dynamic_excel_with_validation.py
# v2+: Dynamic Excel-driven Tkinter data-entry app with adaptive validation
#
# Save as: data_entry_v2_dynamic_excel_with_validation.py
# Requirements:
# pip install ttkbootstrap openpyxl

import os
import re
import warnings
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from ttkbootstrap import Window, Style
from openpyxl import load_workbook, Workbook
from datetime import datetime, date
import pandas as pd
from openpyxl.utils.dataframe import dataframe_to_rows

APP_TITLE = "tEppy's Data Entry (Excel Companion with validation)"

# -------------------------
# Tooltip Helper
# -------------------------
class ToolTip:
    """Simple tooltip for any widget."""
    def __init__(self, widget, text):
        self.widget = widget
        self.text = text
        self.tip_window = None
        widget.bind("<Enter>", self.show_tip)
        widget.bind("<Leave>", self.hide_tip)

    def show_tip(self, event=None):
        if self.tip_window or not self.text:
            return
        x, y, cx, cy = self.widget.bbox("insert")
        x += self.widget.winfo_rootx() + 25
        y += self.widget.winfo_rooty() + 25
        self.tip_window = tw = tk.Toplevel(self.widget)
        tw.wm_overrideredirect(True)  # no window border
        tw.wm_geometry(f"+{x}+{y}")
        label = ttk.Label(
            tw,
            text=self.text,
            justify=tk.LEFT,
            background="#ffffe0",
            relief=tk.SOLID,
            borderwidth=1,
            padding=(6, 3)
        )
        label.pack(ipadx=4)

    def hide_tip(self, event=None):
        tw = self.tip_window
        if tw:
            tw.destroy()
        self.tip_window = None

# --------------------------------------------
# Universal Excel Loader
# --------------------------------------------
def load_any_excel(path: str, app_instance=None):
    """
    Load Excel, binary, or ODS file into an openpyxl Workbook.
    Supports: .xlsx, .xlsm, .xlsb, .xls, .ods
    Returns an openpyxl Workbook instance.
    Optionally takes `app_instance` (DynamicExcelApp) to show in-app popups.
    """
    ext = os.path.splitext(path)[1].lower()
    try:
        if ext in [".xlsx", ".xlsm"]:
            with warnings.catch_warnings(record=True) as w:
                warnings.simplefilter("always")
                wb = load_workbook(path, data_only=True)

            # --- Actively detect embedded images ---
            unsupported_formats = (".wmf", ".emf", ".tiff", ".bmp")
            for ws in wb.worksheets:
                for image in getattr(ws, "_images", []):
                    if any(fmt in str(getattr(image, 'path', '')).lower() for fmt in unsupported_formats):
                        if app_instance:
                            app_instance._show_temp_warning(
                                "⚠️ Workbook contains embedded images (e.g., WMF/EMF).\n"
                                "Avoid opening sheets with images to prevent data loss.",
                                5000
                            )
                        break
            # --- Optional: still catch other openpyxl warnings ---
            for warn in w:
                msg = str(warn.message).lower()
                if "data validation" in msg and app_instance:
                    app_instance._show_temp_warning(
                        "⚠️ Some Excel validations may not load correctly.\n"
                        "Data is safe, but rules will be removed on save.",
                        5000
                    )
                    break

            return wb

        elif ext == ".xlsb":
            df = pd.read_excel(path, engine="pyxlsb")

        elif ext == ".xls":
            df = pd.read_excel(path, engine="xlrd")

        elif ext == ".ods":
            df = pd.read_excel(path, engine="odf")

        else:
            raise ValueError(f"Unsupported file format: {ext}")

        # --- Convert DataFrame to openpyxl Workbook for non-xlsx formats ---
        wb = Workbook()
        ws = wb.active
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
        return wb

    except Exception as e:
        raise RuntimeError(f"Failed to load file ({ext}): {e}")


# -------------------------
# Validation helpers
# -------------------------
EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")

def try_parse_date(value: str):
    """Try multiple common date formats. Return date object if successful, else None."""
    if value is None:
        return None
    s = value.strip()
    if s == "":
        return None
    formats = ["%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y", "%Y/%m/%d"]
    for fmt in formats:
        try:
            return datetime.strptime(s, fmt).date()
        except Exception:
            continue
    # try ISO fallback
    try:
        return datetime.fromisoformat(s).date()
    except Exception:
        return None

def is_numeric(value: str):
    if value is None:
        return False
    s = str(value).strip()
    if s == "":
        return False
    try:
        float(s.replace(",", ""))  # allow commas in thousands
        return True
    except Exception:
        return False

def normalize_numeric(value: str, fmt: str = "integer"):
    """Normalize numeric strings based on desired format."""
    if value is None or str(value).strip() == "":
        return None

    num = float(str(value).replace(",", "").strip())

    if fmt == "decimal":
        return round(num, 2)  # Keep 2 decimal places
    else:
        # Default: treat as integer-like, but preserve float if not whole
        if num.is_integer():
            return int(num)
        else:
            return num


# -------------------------
# Main app
# -------------------------
class DynamicExcelApp:
    def __init__(self, root: Window):
        self.root = root
        self.root.title(APP_TITLE)
        self.root.geometry("1100x720")

        # State
        self.workbook = None
        self.filepath = None
        self.active_sheet_name = None
        self.headers = []
        self.input_entries = []  # list of tk.Entry widgets matching headers
        self.unsaved_changes = False
        self.original_editing_values = {} # NEW: Store original values for uniqueness check exclusion
        self.mode = "add" # Can be "add" or "edit"

        # Inferred validation rules per column: list of dicts
        self.validation_rules = []
        # Reference for theming (bottom-frame needs this)
        self.style = ttk.Style()

        # UI elements
        self._create_menu()
        self._create_toolbar()
        self._create_top_frame()
        self._create_bottom_frame()
        self._create_statusbar()
        self._bind_events()
        self._bind_shortcuts()
        # NOTE: Auto-save var is attached in the `main` function below the class definition

        # Start with a clean app — prompt user to open file
        self._prompt_open_file_on_startup()

    def _bind_shortcuts(self):
        """Bind common Excel-like keyboard shortcuts."""
        self.root.bind("<Control-o>", lambda e: self.open_file())
        self.root.bind("<Control-s>", lambda e: self.save_file())
        self.root.bind("<Control-S>", lambda e: self.save_file_as())  # Shift+S
        self.root.bind("<Control-n>", lambda e: self.clear_input_entries())
        self.root.bind("<Control-d>", lambda e: self._duplicate_selected_row()) 
        self.root.bind("<F2>", lambda e: self.on_tree_double_click(e))
        self.root.bind("<Escape>", lambda e: self.reset_to_add_mode())
        # (Optional) Ctrl+Q to exit
        self.root.bind("<Control-q>", lambda e: self.on_close())
        self.root.bind("<Control-Shift-I>", lambda e: self._insert_blank_row_below_selection())
        self.root.bind("<Control-Shift-D>", lambda e: self.delete_selected_row())

    # -----------------------
    # UI Construction
    # -----------------------
    def _create_menu(self):
        menubar = tk.Menu(self.root)
        file_menu = tk.Menu(menubar, tearoff=0)
        file_menu.add_command(label="Open...", command=self.open_file)
        file_menu.add_command(label="Save", command=self.save_file)
        file_menu.add_command(label="Save As...", command=self.save_file_as)
        file_menu.add_separator()
        file_menu.add_command(label="Exit", command=self.on_close)
        menubar.add_cascade(label="File", menu=file_menu)

        help_menu = tk.Menu(menubar, tearoff=0)
        help_menu.add_command(label="Help & Instructions", command=self._show_help)
        help_menu.add_separator()
        help_menu.add_command(label="About", command=self._show_about)
        menubar.add_cascade(label="Help", menu=help_menu)

        self.root.config(menu=menubar)

    def _create_toolbar(self):
        """Top toolbar: quick actions + sheet selector + Auto-Save + Theme + Help."""
        toolbar = ttk.Frame(self.root, padding=(8, 6))
        toolbar.pack(side=tk.TOP, fill=tk.X)

        # --- Left Section: Core Actions ---
        delete_btn = ttk.Button(toolbar, text="🗑 Delete Row", command=self.delete_selected_row, style="danger.TButton")
        delete_btn.pack(side=tk.LEFT, padx=(0, 10))

        self.add_button = ttk.Button(toolbar, text="➕ Add Row", command=self.add_row_from_inputs, style="success.TButton")
        self.add_button.pack(side=tk.LEFT, padx=(0, 14))

        ttk.Label(toolbar, text="Sheet:", bootstyle="secondary").pack(side=tk.LEFT, padx=(8, 4))
        self.sheet_combo = ttk.Combobox(toolbar, state="readonly", width=28)
        self.sheet_combo.pack(side=tk.LEFT, padx=(0, 8))
        self.sheet_combo.bind("<<ComboboxSelected>>", self.on_sheet_change)

        # --- Center Spacer ---
        spacer = ttk.Label(toolbar, text="")
        spacer.pack(side=tk.LEFT, expand=True)

        # ===========================
        # Right Group (Auto-Save + Theme + Help)
        # ===========================
        right_grp = ttk.Frame(toolbar, padding=(8, 4))
        right_grp.pack(side=tk.RIGHT)
        right_grp.config(relief=tk.GROOVE, borderwidth=1)

        # Auto-Save Checkbox
        self.auto_save_var = tk.BooleanVar(value=False)
        auto_save_chk = ttk.Checkbutton(
            right_grp,
            text="Auto-Save",
            variable=self.auto_save_var,
            style="primary.TCheckbutton"
        )
        auto_save_chk.pack(side=tk.LEFT, padx=(0, 8))

        # Separator (visual divider)
        ttk.Separator(right_grp, orient="vertical").pack(side=tk.LEFT, fill=tk.Y, padx=(4, 10))

        # Theme Selector
        ttk.Label(right_grp, text="Theme:", bootstyle="secondary").pack(side=tk.LEFT, padx=(4, 4))
        self.theme_combo = ttk.Combobox(right_grp, values=Style().theme_names(), state="readonly", width=15)
        self.theme_combo.set(Style().theme_use())
        self.theme_combo.bind("<<ComboboxSelected>>", self.on_theme_change)
        self.theme_combo.pack(side=tk.LEFT, padx=(0, 6))

        # Help Button
        help_btn = ttk.Button(
            right_grp,
            text="❓ Help",
            command=self._show_help,
            style="info.TButton",
            width=8
        )
        help_btn.pack(side=tk.LEFT, padx=(10, 0))
        ToolTip(help_btn, "View usage instructions (from help.txt)")
        self._add_hover_effect(help_btn)

        # --- Apply tooltips and hover globally ---
        ToolTip(delete_btn, "Delete the selected row from sheet and view.")
        ToolTip(self.add_button, "Add a new row using data from input fields.")
        ToolTip(auto_save_chk, "Automatically save after Add/Edit/Delete actions.")
        ToolTip(self.theme_combo, "Switch between ttkbootstrap themes.")
        ToolTip(help_btn, "View usage instructions (from help.txt)")

        for widget in [delete_btn, self.add_button, auto_save_chk, help_btn]:
            self._add_hover_effect(widget)

    def _create_statusbar(self):
        """Modern status bar (Excel-style, right-aligned with color-coded updates)."""
        self.status_var = tk.StringVar(value="Ready")
        self.status_label = ttk.Label(
            self.root,
            textvariable=self.status_var,
            anchor="e",  # right-align text
            padding=(6, 2),
            bootstyle="secondary",
        )
        self.status_label.pack(side=tk.BOTTOM, fill=tk.X)

    def _create_top_frame(self):
        # top frame will contain dynamic input fields (headers -> entries)
        self.top_frame = ttk.Frame(self.root)
        self.top_frame.pack(side=tk.TOP, fill=tk.X, padx=10, pady=(6, 0))

        # We'll put inputs into a scrollable frame in case there are many columns
        self.input_canvas = tk.Canvas(self.top_frame, height=140)
        self.input_canvas.pack(side=tk.TOP, fill=tk.X, expand=True)

        self.input_scrollbar = ttk.Scrollbar(self.top_frame, orient=tk.HORIZONTAL, command=self.input_canvas.xview)
        self.input_scrollbar.pack(side=tk.TOP, fill=tk.X)
        self.input_canvas.configure(xscrollcommand=self.input_scrollbar.set)

        self.inputs_inner = ttk.Frame(self.input_canvas)
        self.input_canvas.create_window((0, 0), window=self.inputs_inner, anchor="nw")
        self.inputs_inner.bind("<Configure>", lambda e: self.input_canvas.configure(scrollregion=self.input_canvas.bbox("all")))

    def _create_bottom_frame(self):
        """
        Bottom section containing the Treeview and the header-based filter row (simplified layout).
        """
        bottom_frame = ttk.Frame(self.root)
        bottom_frame.pack(side=tk.TOP, fill=tk.BOTH, expand=True, padx=10, pady=10)

        # --- Filter Row Frame (no canvas, just simple frame above tree) ---
        self.filter_frame = ttk.Frame(bottom_frame)
        self.filter_frame.pack(side=tk.TOP, fill=tk.X, pady=(0, 3))

        # --- Treeview Frame ---
        tree_frame = ttk.Frame(bottom_frame)
        tree_frame.pack(side=tk.TOP, fill=tk.BOTH, expand=True)

        self.tree = ttk.Treeview(tree_frame, show="headings")
        self.tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        vsb = ttk.Scrollbar(tree_frame, orient="vertical", command=self.tree.yview)
        vsb.pack(side=tk.LEFT, fill=tk.Y)
        self.tree.configure(yscrollcommand=vsb.set)

        hsb = ttk.Scrollbar(tree_frame, orient="horizontal", command=self.tree.xview)
        hsb.pack(side=tk.BOTTOM, fill=tk.X)
        self.tree.configure(xscrollcommand=hsb.set)

        # --- Data caches ---
        self.all_rows = []
        self.filter_entries = []
        self.column_ids = []

        # When Treeview resizes, adjust entry widths
        self.tree.bind("<Configure>", lambda e: self._adjust_filter_widths())
        self.tree.bind("<ButtonRelease-1>", lambda e: self._adjust_filter_widths())  # for resize drag

    def _on_tree_scroll(self, *args):
        """
        Sync horizontal scroll between Treeview and filter canvas.
        """
        self.tree.xview(*args)
        self.filter_canvas.xview(*args)
        self._sync_filter_positions()

    def _create_filter_row(self):
        """
        Creates one filter entry per column directly above the Treeview.
        """
        # Clear previous filters
        for w in self.filter_frame.winfo_children():
            w.destroy()
        self.filter_entries.clear()
        self.column_ids = list(self.tree["columns"])

        if not self.column_ids:
            return

        # Create one Entry per column
        for col_id in self.column_ids:
            entry = ttk.Entry(self.filter_frame)
            entry.pack(side=tk.LEFT, padx=1, fill=tk.X, expand=True)
            entry.insert(0, "")
            entry.bind("<KeyRelease>", lambda e: self._apply_filters())
            ToolTip(entry, f"Filter '{self.tree.heading(col_id, 'text')}'")
            self.filter_entries.append(entry)

        self.root.after(100, self._adjust_filter_widths)

    def _adjust_filter_widths(self):
        """
        Match filter entry widths to Treeview column widths.
        """
        if not self.filter_entries or not self.column_ids:
            return

        # Get total width of the tree
        total_width = sum(int(self.tree.column(col, "width")) for col in self.column_ids)
        self.filter_frame.update_idletasks()

        for i, col_id in enumerate(self.column_ids):
            try:
                width = int(self.tree.column(col_id, "width"))
            except tk.TclError:
                width = 100
            self.filter_entries[i].config(width=max(8, width // 10))

    def _sync_filter_positions(self):
        """
        Align filter entries under each Treeview column, using cumulative column widths.
        Works even when no data rows exist.
        """
        if not self.filter_entries or not self.column_ids:
            return

        # Ensure Canvas is visible and ready
        self.filter_canvas.update_idletasks()

        x_offset = 0
        total_width = 0

        for i, col_id in enumerate(self.column_ids):
            try:
                width = int(self.tree.column(col_id, 'width'))
            except tk.TclError:
                width = 120  # fallback width

            entry = self.filter_entries[i]
            entry.place(x=x_offset, y=2, width=width, height=26)
            x_offset += width
            total_width += width

        # Adjust scroll region so entries can be seen/scrolled properly
        self.filter_canvas.configure(scrollregion=(0, 0, total_width, 30))
        self.filter_inner.update_idletasks()
        self.filter_canvas.update_idletasks()

    def _apply_filters(self):
        """
        Simple per-column substring filtering.
        """
        if not self.all_rows:
            return

        filters = [f.get().strip().lower() for f in self.filter_entries]
        if all(f == "" for f in filters):
            self._reload_tree_from_cache()
            return

        filtered = []
        for row in self.all_rows:
            if all(
                (f in str(row[i]).lower() if f else True)
                for i, f in enumerate(filters)
            ):
                filtered.append(row)

        self._reload_tree_from_cache(filtered)

    def _reload_tree_from_cache(self, rows=None):
        """
        Reload Treeview data from cache.
        """
        self.tree.delete(*self.tree.get_children())
        display_rows = rows if rows is not None else self.all_rows

        for row in display_rows:
            row_extended = list(row) + [""] * (len(self.headers) - len(row))
            self.tree.insert("", tk.END, values=row_extended)

    def _clear_filters(self):
        """
        Clear all filters and restore full dataset view.
        """
        if not hasattr(self, "filter_entries"):
            return

        for entry in self.filter_entries:
            entry.delete(0, tk.END)

        self._reload_tree_from_cache()
        self._update_status("Filters cleared. Showing all data.", "success")

    def _bind_events(self):
        # Save prompt on close
        self.root.protocol("WM_DELETE_WINDOW", self.on_close)
        # Bind double-click editing
        self.tree.bind("<Double-1>", self.on_tree_double_click)

    # -----------------------
    # File and Sheet Handling
    # -----------------------
    def _prompt_open_file_on_startup(self):
        answer = messagebox.askyesno("Open file", "Is your template ready for loading?")
        if answer:
            self.open_file()
        else:
            self.status_var.set("Ready. Use File -> Open or Ctrl+O to open a spreadsheet.")

    def open_file(self):
        filetypes = [
            ("All supported files", "*.xlsx *.xlsm *.xlsb *.xls *.ods"),
            ("Excel files", "*.xlsx;*.xlsm;*.xlsb;*.xls"),
            ("LibreOffice files", "*.ods"),
        ]
        path = filedialog.askopenfilename(title="Open spreadsheet", filetypes=filetypes)
        if not path:
            return

        try:
            wb = load_any_excel(path)
        except Exception as e:
            messagebox.showerror("Error", f"Failed to open file:\n{e}")
            return

        self.workbook = wb
        self.filepath = path
        self.active_sheet_name = self.workbook.active.title
        self._populate_sheet_selector()
        self._load_active_sheet()
        self.unsaved_changes = False

        # Warn if it’s a non-.xlsx file
        ext = os.path.splitext(path)[1].lower()
        if ext not in [".xlsx", ".xlsm"]:
            self._update_status(f"Opened {os.path.basename(path)} (converted to in-memory .xlsx)")
            messagebox.showinfo(
                "Format Notice",
                "This file was opened from a non-.xlsx format.\n\n"
                "It will be saved as .xlsx when you save changes."
            )
        else:
            self._update_status(f"Opened: {os.path.basename(path)}")

    def _insert_blank_row_below_selection(self):
        """Insert a blank row below the selected row (Excel-style)."""
        if not self.workbook or not self.active_sheet_name:
            messagebox.showwarning("No file", "Open an .xlsx file first.")
            return

        selected_item = self.tree.focus()
        sheet = self.workbook[self.active_sheet_name]

        if selected_item:
            current_index = self.tree.index(selected_item)
            excel_row_index = current_index + 3  # +2 header, +1 for below current
            insert_index = current_index + 1
        else:
            # If nothing selected, insert at top of data area
            excel_row_index = 2
            insert_index = 0

        try:
            sheet.insert_rows(excel_row_index, 1)
            blank_values = ["" for _ in self.headers]
            new_item = self.tree.insert("", insert_index, values=blank_values)

            # --- FIXED VISUAL SELECTION BEHAVIOR ---
            self.tree.selection_remove(self.tree.selection())
            self.tree.selection_set(new_item)
            self.tree.focus(new_item)
            self.tree.see(new_item)

            # Flash green to indicate insertion
            self._flash_tree_row(new_item, color="#ccffcc", duration=700)

            self.unsaved_changes = True
            self._update_status(f"Inserted blank row at Excel row {excel_row_index}.")

            # Auto-save if enabled
            if hasattr(self, 'auto_save_var') and self.auto_save_var.get():
                if self.save_file():
                    self._update_status("Auto-saved after inserting row ✅")

        except Exception as e:
            messagebox.showerror("Error", f"Failed to insert row:\n{e}")

    def _flash_tree_row(self, item_id, color="#ccffcc", duration=800):
        """Temporarily flash a row background for visual feedback."""
        try:
            self.tree.tag_configure("flash", background=color)
            self.tree.item(item_id, tags=("flash",))
            self.root.after(duration, lambda: self.tree.item(item_id, tags=()))
        except Exception:
            pass

    def _populate_sheet_selector(self):
        if not self.workbook:
            self.sheet_combo["values"] = []
            return
        names = self.workbook.sheetnames
        self.sheet_combo["values"] = names
        if self.active_sheet_name in names:
            self.sheet_combo.set(self.active_sheet_name)
        else:
            self.sheet_combo.set(names[0])
            self.active_sheet_name = names[0]

    def on_sheet_change(self, event=None):
        if not self.workbook:
            return
        new_sheet = self.sheet_combo.get()
        if new_sheet == self.active_sheet_name:
            return

        if self.unsaved_changes:
            res = messagebox.askyesnocancel("Unsaved changes", "You have unsaved changes. Save before switching sheets?")
            if res is None:
                # Cancel switching
                self.sheet_combo.set(self.active_sheet_name)
                return
            if res:  # Yes -> save
                if not self.save_file():
                    # Save failed or cancelled; cancel switching
                    self.sheet_combo.set(self.active_sheet_name)
                    return

        self.active_sheet_name = new_sheet
        self._load_active_sheet()

    def _load_active_sheet(self):
        """Read headers and rows from active sheet, rebuild inputs, treeview, and filters."""
        if not self.workbook or not self.active_sheet_name:
            return

        sheet = self.workbook[self.active_sheet_name]

        # --- Load headers (same logic as before) ---
        headers = []
        header_row_idx = None
        for r in sheet.iter_rows(min_row=1, max_row=5):
            values = [cell.value for cell in r]
            if any(v is not None and str(v).strip() != "" for v in values):
                header_row_idx = r[0].row
                headers = [
                    str(cell.value).strip() if cell.value is not None and str(cell.value).strip() != "" else None
                    for cell in r
                ]
                break

        if not headers:
            max_col = sheet.max_column or 1
            headers = [None] * max_col
            header_row_idx = 1

        self.headers = [h if h else f"Column {i+1}" for i, h in enumerate(headers)]

        # Infer validation rules and build inputs
        self.validation_rules = self._infer_validation_rules(self.headers)
        self._build_input_fields(self.headers)

        # --- Load data rows into Treeview + cache ---
        self._clear_treeview()
        cols = [f"c{i}" for i in range(len(self.headers))]
        self.tree["columns"] = cols
        for i, h in enumerate(self.headers):
            self.tree.heading(cols[i], text=h, anchor=tk.W)
            self.tree.column(cols[i], width=160, anchor=tk.W)

        start_row = header_row_idx + 1
        rows = []
        for r in sheet.iter_rows(min_row=start_row, max_row=sheet.max_row, max_col=len(self.headers)):
            rowvals = [cell.value if cell.value is not None else "" for cell in r]
            if all(v == "" or v is None for v in rowvals):
                continue
            rows.append(rowvals)

        self.all_rows = rows  # cache for filtering
        for row in rows:
            row_extended = list(row) + [""] * (len(self.headers) - len(row))
            self.tree.insert("", tk.END, values=row_extended)

        # --- Finally: build the filter row now that we know the headers ---
        self._create_filter_row()

    def _duplicate_selected_row(self):
        """Duplicate the currently selected row (inserted right below it), auto-incrementing ID-like fields."""
        if not self.workbook or not self.active_sheet_name:
            messagebox.showwarning("No file", "Open an .xlsx file first.")
            return

        selected_item = self.tree.focus()
        if not selected_item:
            messagebox.showinfo("No selection", "Please select a row to duplicate.")
            return

        try:
            # Get current row data
            values = list(self.tree.item(selected_item, "values"))
            sheet = self.workbook[self.active_sheet_name]
            headers = self.headers

            # Define which headers should be treated as ID fields
            id_keywords = ("id", "no", "code", "ref", "reference")

            # Auto-increment ID-like fields when possible
            new_values = []
            for h, v in zip(headers, values):
                if not h:
                    new_values.append(v)
                    continue

                h_lower = str(h).lower()
                v_str = str(v).strip()

                if any(k in h_lower for k in id_keywords) and v_str != "":
                    # Try to find trailing numeric pattern (e.g. "ID001" -> "ID002")
                    import re
                    match = re.search(r"(\d+)$", v_str)
                    if match:
                        prefix = v_str[:match.start()]
                        num = match.group(1)
                        new_num = str(int(num) + 1).zfill(len(num))  # preserve zero-padding
                        new_values.append(prefix + new_num)
                    elif v_str.isdigit():
                        new_values.append(str(int(v_str) + 1))
                    else:
                        new_values.append(v_str)
                else:
                    new_values.append(v)

            # Determine where to insert (below the current row)
            current_index = self.tree.index(selected_item)
            excel_row_index = current_index + 3  # +2 header +1 for below current

            # Insert new row in workbook and copy values
            sheet.insert_rows(excel_row_index, 1)
            for col_index, val in enumerate(new_values, start=1):
                sheet.cell(row=excel_row_index, column=col_index).value = val

            # Insert new row visually in Treeview (below current)
            new_item = self.tree.insert("", current_index + 1, values=new_values)

            # --- FIXED SELECTION BEHAVIOR ---
            # Clear previous selection and move highlight to the new row
            self.tree.selection_remove(self.tree.selection())
            self.tree.selection_set(new_item)
            self.tree.focus(new_item)   # Force the focus to match the visual highlight
            self.tree.see(new_item)

            # Flash green to show duplication success
            self._flash_tree_row(new_item, color="#ccffcc", duration=700)

            self.unsaved_changes = True
            self._update_status(f"Duplicated row {current_index + 1} with auto-incremented IDs.")

            # Auto-save if enabled
            if hasattr(self, 'auto_save_var') and self.auto_save_var.get():
                if self.save_file():
                    self._update_status("Auto-saved after duplicating row ✅")

        except Exception as e:
            messagebox.showerror("Error", f"Failed to duplicate row:\n{e}")

    # -----------------------
    # Validation rule inference (UPDATED)
    # -----------------------
    def _infer_validation_rules(self, headers):
        """
        Heuristic inference of validation rules from header names, including numeric formatting awareness.
        Returns a list of dicts: {'name': str, 'type': 'text'|..., 'format': 'decimal'|'integer'|None, ...}
        """
        rules = []
        numeric_keywords = ("qty", "quantity", "number", "count", "age")
        decimal_keywords = ("amount", "price", "rate", "total", "cost", "balance", "value")

        for h in headers:
            h_lower = h.lower() if h else ""
            
            # --- Default states ---
            val_type = "text"
            num_format = None
            is_required_default = True
            duplicate_policy_default = "none"

            # Determine Type & Formatting
            if any(k in h_lower for k in decimal_keywords):
                val_type = "numeric"
                num_format = "decimal"
            elif any(k in h_lower for k in numeric_keywords):
                val_type = "numeric"
                num_format = "integer"
            elif "date" in h_lower:
                val_type = "date"
            elif "email" in h_lower:
                val_type = "email"

            # Duplicate/required logic
            if "optional" in h_lower:
                is_required_default = False
            if "id" in h_lower or "code" in h_lower:
                duplicate_policy_default = "strict"
                is_required_default = False

            rule = {
                "name": h,
                "type": val_type,
                "format": num_format,  # 👈 new key
                "required_var": tk.BooleanVar(value=is_required_default),
                "duplicate_var": tk.StringVar(value=duplicate_policy_default),
                "required": is_required_default,
                "duplicate_policy": duplicate_policy_default
            }

            rules.append(rule)

        return rules

    # -----------------------
    # Inputs builder & validation (UPDATED)
    # -----------------------
    def _clear_inputs_area(self):
        for w in self.inputs_inner.winfo_children():
            w.destroy()
        self.input_entries.clear()

    def _build_input_fields(self, headers):
        """
        Create horizontal layout of label+entry with QoL validation controls per header.
        (Updated duplicate policy labels for clarity)
        """
        self._clear_inputs_area()

        for idx, header in enumerate(headers):
            rule = self.validation_rules[idx]

            col_frame = ttk.Frame(self.inputs_inner)
            col_frame.grid(row=0, column=idx, padx=6, pady=4)

            lbl = ttk.Label(col_frame, text=header, width=20, anchor="center")
            lbl.pack(side=tk.TOP, fill=tk.X)

            ent = tk.Entry(col_frame, width=20)
            ent.pack(side=tk.TOP, pady=(6, 0))
            ent.bind("<Return>", lambda e, i=idx: self._on_enter_pressed(e, i))
            ent.bind("<Tab>", lambda e, i=idx: (self._on_enter_pressed(e, i), "break")[1])
            self.input_entries.append(ent)

            error_var = tk.StringVar(value="")
            error_lbl = ttk.Label(col_frame, textvariable=error_var, foreground="red", anchor="center")
            error_lbl.pack(side=tk.TOP, fill=tk.X)
            ent.error_var = error_var

            # QoL Controls Frame
            control_frame = ttk.Frame(col_frame)
            control_frame.pack(side=tk.TOP, fill=tk.X, pady=(5, 0))

            # Required Checkbox
            req_chk = ttk.Checkbutton(
                control_frame,
                text="Required",
                variable=rule['required_var'],
                command=lambda r=rule: self._update_validation_state(r)
            )
            req_chk.pack(anchor=tk.W)

            # Duplicate Policy Label
            dup_lbl = ttk.Label(control_frame, text="Duplicate Policy:")
            dup_lbl.pack(anchor=tk.W, pady=(2, 0))

            # Duplicate Options
            ttk.Radiobutton(control_frame, text="None", variable=rule['duplicate_var'], value="none",
                            command=lambda r=rule: self._update_validation_state(r)).pack(anchor=tk.W, padx=10)
            ttk.Radiobutton(control_frame, text="Warn (Allow Duplicates)", variable=rule['duplicate_var'], value="warn",
                            command=lambda r=rule: self._update_validation_state(r)).pack(anchor=tk.W, padx=10)
            ttk.Radiobutton(control_frame, text="No Duplicates (Strict)", variable=rule['duplicate_var'], value="strict",
                            command=lambda r=rule: self._update_validation_state(r)).pack(anchor=tk.W, padx=10)

        self.root.after(100, lambda: self.input_canvas.configure(scrollregion=self.input_canvas.bbox("all")))
        self.reset_to_add_mode()

    def _update_validation_state(self, rule):
        """Updates the internal validation rule dictionary from the user's QoL controls."""
        
        # Read the current Tk variable states
        new_required = rule['required_var'].get()
        new_duplicate_policy = rule['duplicate_var'].get()
        
        # Update the rule dictionary used by validate_inputs
        rule["required"] = new_required
        rule["duplicate_policy"] = new_duplicate_policy
        
        self._update_status(f"Validation policy updated for '{rule['name']}'. Required: {new_required}, Duplicate: {new_duplicate_policy}")

    def _on_enter_pressed(self, event, idx):
        # If Enter pressed in last field, add row; else focus next.
        if idx == len(self.input_entries) - 1:
            if self.mode == "add":
                self.add_row_from_inputs()
            elif self.mode == "edit":
                 self.update_row_from_inputs()
        else:
            self.input_entries[idx + 1].focus_set()

    def _get_existing_column_data(self, col_index):
        """
        NEW HELPER: Retrieves all non-empty, normalized values from a specific column 
        in the active sheet (data rows only). Returns a set for fast lookup.
        """
        if not self.workbook or not self.active_sheet_name:
            return set()
            
        sheet = self.workbook[self.active_sheet_name]
        
        # We assume the data starts at row 2 (after the header row)
        existing_values = set()
        
        # Iterate over all rows starting from the data rows (row 2 or higher)
        for row_idx in range(2, sheet.max_row + 1): 
            cell_value = sheet.cell(row=row_idx, column=col_index).value
            if cell_value is not None:
                # Normalize the value (strip whitespace, convert to string) for case-insensitive comparison
                normalized_value = str(cell_value).strip()
                if normalized_value:
                    existing_values.add(normalized_value.lower()) 
                    
        return existing_values
        
    def validate_inputs(self):
        """
        Validate all input entries. Applies background color and error labels for feedback.
        Returns (is_valid: bool, strict_messages: list, warning_messages: list, normalized_values: list)
        """
        normalized = []
        strict_messages = []
        warning_messages = [] 
        is_valid = True
        
        is_edit_mode = self.mode == "edit"

        for i, entry in enumerate(self.input_entries):
            val = entry.get()
            rule = self.validation_rules[i]
            col_name = rule["name"]
            val_type = rule["type"]
            # --- CRITICAL: USE UPDATED STATE KEYS ---
            required = rule["required"] 
            duplicate_policy = rule["duplicate_policy"]
            
            # Reset feedback for this entry
            entry.config(bg="white")
            if hasattr(entry, 'error_var'):
                entry.error_var.set("")
            
            val_stripped = val.strip()
            
            # 1. Required check
            if required and not val_stripped:
                is_valid = False
                strict_messages.append(f"❌ {col_name}: Required.") # <-- CORRECTED
                normalized.append(None)
                entry.config(bg="#fbb") # Light red background for error
                if hasattr(entry, 'error_var'):
                    entry.error_var.set("Required")
                continue # Move to next field

            # Skip type validation for optional, empty fields
            if not val_stripped and not required:
                 normalized.append(None)
                 continue
            
            # 2. Uniqueness Check (MODIFIED LOGIC)
            if duplicate_policy in ("strict", "warn"):
                col_index = i + 1
                existing_data = self._get_existing_column_data(col_index=col_index)
                
                # Exclusion Logic for Edit Mode
                is_original_value = False
                if is_edit_mode:
                    original_val = self.original_editing_values.get(col_index, "__NO_MATCH__")
                    # Check against the normalized original value
                    if val_stripped.lower() == original_val:
                        is_original_value = True
                        
                # Only check for duplication if it's NOT the original value
                if not is_original_value and val_stripped.lower() in existing_data:
                    # DUPLICATE VIOLATION FOUND
                    if duplicate_policy == "strict":
                        # Treat as an error
                        is_valid = False
                        strict_messages.append(f"❌ {col_name}: Duplicate value found (Strict Policy).")
                        entry.config(bg="#fbb")
                        if hasattr(entry, 'error_var'):
                            entry.error_var.set("Duplicate (Strict)")
                        normalized.append(val_stripped)
                        continue # Stop further checks for this field
                    
                    elif duplicate_policy == "warn":
                        # Treat as a warning, but keep is_valid = True
                        warning_messages.append(f"⚠️ {col_name}: Possible duplicate value found.")
                        entry.config(bg="#ffdd99") # Yellow/orange background for warning
                        if hasattr(entry, 'error_var'):
                            entry.error_var.set("Possible Duplicate")
                        # DO NOT continue, proceed to type check
            
            # 3. Type-specific validation and normalization
            try:
                if val_type == "text":
                    cleaned = re.sub(r"\s+", " ", val_stripped)
                    normalized.append(cleaned if cleaned else None)
                    
                elif val_type == "numeric":
                    if not is_numeric(val):
                        raise ValueError("Invalid number")
                    normalized.append(normalize_numeric(val, fmt=rule.get("format", "integer")))

                elif val_type == "date":
                    date_obj = try_parse_date(val)
                    if date_obj is None:
                        raise ValueError("Invalid date (Try YYYY-MM-DD)")
                    normalized.append(date_obj) 

                elif val_type == "email":
                    if not EMAIL_RE.match(val_stripped):
                        raise ValueError("Invalid email format")
                    normalized.append(val_stripped)

                else:
                    normalized.append(val_stripped)

            except ValueError as e:
                is_valid = False
                strict_messages.append(f"❌ {col_name}: {e}") # <-- CORRECTED
                entry.config(bg="#fbb")
                if hasattr(entry, 'error_var'):
                    entry.error_var.set(str(e))
                normalized.append(val_stripped)

        return is_valid, strict_messages, warning_messages, normalized

    def clear_input_entries(self):
        for ent in self.input_entries:
            ent.delete(0, tk.END)
            # Reset background and error label
            ent.config(bg="white")
            if hasattr(ent, 'error_var'):
                ent.error_var.set("")

    # -----------------------
    # Treeview population
    # -----------------------
    def _clear_treeview(self):
        for col in self.tree["columns"]:
            self.tree.heading(col, text="")
        self.tree.delete(*self.tree.get_children())
        self.tree["columns"] = ()

    def _load_treeview_rows(self, sheet, header_row_idx):
        self._clear_treeview()
        # Treeview columns -> headers
        cols = [f"c{i}" for i in range(len(self.headers))]
        self.tree["columns"] = cols
        for i, h in enumerate(self.headers):
            self.tree.heading(cols[i], text=h, anchor=tk.W)
            # set column width heuristically
            self.tree.column(cols[i], width=160, anchor=tk.W)

        # Read rows starting after header_row_idx
        start_row = header_row_idx + 1
        rows = []
        for r in sheet.iter_rows(min_row=start_row, max_row=sheet.max_row, max_col=len(self.headers)):
            rowvals = [ (cell.value if cell.value is not None else "") for cell in r ]
            # If the row is entirely empty, skip
            if all((v == "" or v is None) for v in rowvals):
                continue
            rows.append(rowvals)

        for row in rows:
            # Ensure row length matches headers
            row_extended = list(row) + [""] * (len(self.headers) - len(row))
            self.tree.insert("", tk.END, values=row_extended)

    # -----------------------
    # Add / Edit / Delete Row
    # -----------------------
    def add_row_from_inputs(self):
        """Append a new row after validating input fields."""
        if not self.workbook or not self.active_sheet_name:
            messagebox.showwarning("No file", "Open an .xlsx file first.")
            return

        # Validate inputs # --- CRITICAL CHANGE: Capture warnings list ---
        is_valid, strict_messages, warning_messages, normalized = self.validate_inputs()

        # 1. Handle Strict Validation Failure
        if not is_valid:
            self._update_status(f"Strict validation failed on {len(strict_messages)} field(s).", "error")
            return
            
        # 2. Handle Duplicate Warnings and Prompt User
        if warning_messages:
            warning_text = "\n".join(warning_messages)
            prompt = (
                "Potential duplicate entries were detected:\n\n"
                f"{warning_text}\n\n"
                "Proceed with adding this record?"
            )
            res = messagebox.askyesno("Possible Duplicate Detected", prompt, icon='warning')
            
            if not res:
                self._update_status("Addition cancelled due to duplicate warning.", "warning")
                return # User chose not to proceed

        sheet = self.workbook[self.active_sheet_name]
        append_row_idx = sheet.max_row + 1  
        
        # --- Write values into workbook ---
        for col_index, val in enumerate(normalized, start=1):
            cell = sheet.cell(row=append_row_idx, column=col_index)
            cell.value = val 

        # Reflect in UI: Treeview needs string representation for display
        display_row = []
        for val in normalized:
            if val is None:
                display_row.append("")
            elif isinstance(val, date):
                # Use strftime to format the date as YYYY-MM-DD
                display_row.append(val.strftime("%Y-%m-%d"))
            elif isinstance(val, (int, float)):
                 display_row.append(str(val)) # Convert number to string
            elif isinstance(val, float) and rule.get("format") == "decimal":
                display_row.append(f"{val:.2f}")
            else:
                display_row.append(str(val))

        # Highlight newly-added row
        new_item_id = self.tree.insert("", tk.END, values=display_row)
        self.tree.selection_remove(self.tree.selection())  
        self.tree.selection_set(new_item_id)
        self.tree.see(new_item_id)

        self.unsaved_changes = True
        self._update_status(f"Added new row to '{self.active_sheet_name}'.")

        # Auto-save if enabled
        if hasattr(self, 'auto_save_var') and self.auto_save_var.get():
            if self.save_file():
                self._update_status(f"Auto-saved after adding new row ✅", "success")

        # Reset form
        self.clear_input_entries()
        if self.input_entries:
            self.input_entries[0].focus_set()

    def on_tree_double_click(self, event):
        """Enable editing mode by loading selected row data into input fields."""
        selected_item = self.tree.focus()
        if not selected_item:
            return

        values = self.tree.item(selected_item, "values")
        if not values:
            return

        # Clear previous error feedback before populating new data
        self.clear_input_entries()

        # Populate inputs
        for entry, value in zip(self.input_entries, values):
            entry.delete(0, tk.END)
            entry.insert(0, value)

        self.editing_item = selected_item
        self.mode = "edit"

        # --- NEW: Store original values for uniqueness check exclusion ---
        self.original_editing_values = {}
        for i, val in enumerate(values):
            # Store the normalized (lowercase, stripped) value of the original cell data
            self.original_editing_values[i + 1] = str(val).strip().lower() 
            
        # Change button text + command
        self.add_button.config(text="Update Row", command=self.update_row_from_inputs, style="warning.TButton")

        # Temporarily unbind <Return> from its original 'focus next' logic
        for entry in self.input_entries:
            entry.unbind("<Return>")
            # Bind to a simple update call
            entry.bind("<Return>", lambda e: self.update_row_from_inputs())

        # Focus first field
        if self.input_entries:
            self.input_entries[0].focus_set()

        self._update_status("Editing existing row...", "warning", duration=0)

    def update_row_from_inputs(self):
        """Update selected Treeview row and workbook entry."""
        if not hasattr(self, "editing_item") or not self.editing_item:
            messagebox.showinfo("No selection", "No row selected for editing.")
            return

        # --- CRITICAL CHANGE: Capture warnings list ---
        is_valid, strict_messages, warning_messages, normalized = self.validate_inputs()
        
        # 1. Handle Strict Validation Failure
        if not is_valid:
            self._update_status(f"Strict validation failed on {len(strict_messages)} fields. See red fields.", "error", duration=0)
            return

        # 2. Handle Duplicate Warnings and Prompt User
        if warning_messages:
            warning_text = "\n".join(warning_messages)
            prompt = (
                "The following potential duplicate entries were detected:\n\n"
                f"{warning_text}\n\n"
                "Do you still want to update this record?"
            )
            res = messagebox.askyesno("Possible Duplicate Detected", prompt, icon='warning')
            
            if not res:
                self._update_status("Update cancelled due to duplicate warning.", "warning")
                return # User chose not to proceed

        # Update Treeview - same conversion to display strings as in add_row
        display_row = []
        for val in normalized:
            if val is None:
                display_row.append("")
            elif isinstance(val, date):
                # Use strftime to format the date as YYYY-MM-DD, removing any time component
                display_row.append(val.strftime("%Y-%m-%d"))
            elif isinstance(val, (int, float)):
                display_row.append(str(val)) 
            elif isinstance(val, float) and rule.get("format") == "decimal":
                display_row.append(f"{val:.2f}")
            else:
                display_row.append(str(val))

        self.tree.item(self.editing_item, values=display_row) 

        # Update workbook - use normalized values (ready for openpyxl)
        sheet = self.workbook[self.active_sheet_name]
        # +2 for header row (assuming header is in row 1)
        row_index = self.tree.index(self.editing_item) + 2 
        
        for col_index, val in enumerate(normalized, start=1):
            cell = sheet.cell(row=row_index, column=col_index)
            cell.value = val 

        self.unsaved_changes = True
        self._update_status(f"Updated row {row_index - 1} successfully." "success")

        # Re-highlight the updated row
        self.tree.selection_set(self.editing_item)
        self.tree.see(self.editing_item)

        # Auto-save if enabled
        if hasattr(self, 'auto_save_var') and self.auto_save_var.get():
            if self.save_file():
                self._update_status("Auto-saved after editing row ✅", "success", duration=0)

        # Reset UI and rebind Enter key for Add mode
        self.reset_to_add_mode()

    def _flash_tree_row(self, item_id, color="#ccffcc", duration=800):
        """Temporarily flash a row background for visual feedback."""
        try:
            # Create a unique tag name for the flash to avoid clobbering existing tags
            tag_name = f"flash_{item_id}"
            self.tree.tag_configure(tag_name, background=color)
            # Apply tag
            self.tree.item(item_id, tags=(tag_name,))
            # Remove tag after duration
            self.root.after(duration, lambda: self.tree.item(item_id, tags=()))
        except Exception:
            # Fail silently for robustness
            pass

    def delete_selected_row(self):
        """Deletes the selected row from both the Treeview and the workbook with clear visual feedback."""
        if not self.workbook or not self.active_sheet_name:
            messagebox.showwarning("No file", "Open an .xlsx file first.")
            return

        selected_item = self.tree.focus()
        if not selected_item:
            messagebox.showinfo("No selection", "Please select a row to delete.")
            return

        confirm = messagebox.askyesno(
            "Confirm Deletion",
            "Are you sure you want to delete the selected row?\n\n"
            "This action cannot be undone unless you reload the file."
        )
        if not confirm:
            return

        # Flash red before deleting, then perform delete after short delay
        try:
            self._flash_tree_row(selected_item, color="#ffcccc", duration=300)
            self.root.after(300, lambda: self._delete_row_after_flash(selected_item))
        except Exception:
            # Fallback if animation fails
            self._delete_row_after_flash(selected_item)

    def _delete_row_after_flash(self, selected_item):
        """Helper to perform deletion in workbook and Treeview after flash animation."""
        try:
            sheet = self.workbook[self.active_sheet_name]
            excel_row_index = self.tree.index(selected_item) + 2  # +2 for header

            # Delete from workbook and Treeview
            sheet.delete_rows(excel_row_index, 1)
            next_item = self.tree.next(selected_item)
            prev_item = self.tree.prev(selected_item)
            self.tree.delete(selected_item)

            self.unsaved_changes = True
            self._update_status(f"Deleted row {excel_row_index - 1} from '{self.active_sheet_name}'.", "success")

            # --- FIXED SELECTION BEHAVIOR ---
            # Automatically select next or previous row for clarity
            if next_item:
                self.tree.selection_set(next_item)
                self.tree.focus(next_item)
                self.tree.see(next_item)
            elif prev_item:
                self.tree.selection_set(prev_item)
                self.tree.focus(prev_item)
                self.tree.see(prev_item)
            else:
                self.tree.selection_remove(self.tree.selection())
                self.tree.focus("")  # Clear focus if no rows left

            # Reset to Add mode if deleted item was being edited
            if hasattr(self, "editing_item") and self.editing_item == selected_item:
                self.reset_to_add_mode()

            # Auto-save if enabled
            if hasattr(self, 'auto_save_var') and self.auto_save_var.get():
                if self.save_file():
                    self._update_status("Auto-saved after deleting row ✅", "success")

        except Exception as e:
            messagebox.showerror("Error", f"Failed to delete row:\n{e}")

    def reset_to_add_mode(self):
        """Helper to switch back to the default 'Add Row' state."""
        self.clear_input_entries()
        self.add_button.config(text="➕ Add Row", command=self.add_row_from_inputs, style="success.TButton")
        
        # Re-bind Enter key for Add mode
        for idx, entry in enumerate(self.input_entries):
            entry.unbind("<Return>")
            entry.unbind("<Tab>")
            entry.bind("<Return>", lambda e, i=idx: self._on_enter_pressed(e, i))
            entry.bind("<Tab>", lambda e, i=idx: (self._on_enter_pressed(e, i), "break")[1])
        
        self.mode = "add"
        self.editing_item = None
        self.original_editing_values = {} # Clear stored original values
        if self.input_entries:
            self.input_entries[0].focus_set()

    # -----------------------
    # Save Operations
    # -----------------------
    def save_file(self):
        if not self.workbook:
            messagebox.showwarning("No file", "No workbook is currently open.")
            return False

        if not self.filepath:
            return self.save_file_as()

        try:
            self.workbook.save(self.filepath)
            self.unsaved_changes = False
            self._update_status(f"Saved: {os.path.basename(self.filepath)}", "success")
            return True
        except Exception as e:
            messagebox.showerror("Save error", f"Failed to save workbook:\n{e}")
            return False

    def save_file_as(self):
        filetypes = [("Excel files", "*.xlsx")]
        path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=filetypes)
        if not path:
            return False
        self.filepath = path
        try:
            self.workbook.save(self.filepath)
            self.unsaved_changes = False
            self._update_status(f"Saved as: {os.path.basename(self.filepath)}", "success")
            return True
        except Exception as e:
            messagebox.showerror("Save error", f"Failed to save workbook:\n{e}")
            return False

    # -----------------------
    # Theme handling
    # -----------------------
    def on_theme_change(self, event=None):
        theme_name = self.theme_combo.get()
        try:
            Style().theme_use(theme_name)
            self._update_status(f"Theme changed to {theme_name}", "success")
        except Exception as e:
            messagebox.showerror("Theme error", f"Cannot set theme {theme_name}:\n{e}")

    # -----------------------
    # Helpers & closing
    # -----------------------
    def _update_status(self, message, level="info", duration=5000):
        """
        Update status bar message with color-coded feedback and fade-out.
        level: 'info', 'success', 'warning', 'error'
        duration: milliseconds before fade-out
        """
        colors = {
            "info": "#f8f9fa",      # light gray
            "success": "#d1e7dd",   # green tint
            "warning": "#fff3cd",   # yellow tint
            "error": "#f8d7da",     # red tint
        }
        fg_colors = {
            "info": "#333333",
            "success": "#0f5132",
            "warning": "#664d03",
            "error": "#842029",
        }

        self.status_var.set(message)
        bg = colors.get(level, colors["info"])
        fg = fg_colors.get(level, fg_colors["info"])
        self.status_label.configure(background=bg, foreground=fg)

        # --- Fade-out after duration ---
        if duration > 0:
            self.root.after(duration, lambda: self._fade_status())

    def _show_temp_warning(self, message, duration=5000):
        """
        Display a non-interactive popup warning that auto-closes after a few seconds.
        """
        popup = tk.Toplevel(self.root)
        popup.title("⚠️ Warning")
        popup.geometry("400x100+{}+{}".format(self.root.winfo_rootx() + 150, self.root.winfo_rooty() + 100))
        popup.configure(bg="#fff4e5")  # light amber tone
        popup.attributes("-topmost", True)
        popup.resizable(False, False)

        # Remove title bar buttons (platform-dependent)
        try:
            popup.overrideredirect(True)
        except Exception:
            pass

        msg = ttk.Label(
            popup,
            text=message,
            wraplength=380,
            justify="center",
            foreground="#8a6d3b",
            background="#fff4e5",
            font=("Segoe UI", 10, "bold")
        )
        msg.pack(expand=True, fill=tk.BOTH, padx=10, pady=10)

        # Auto-destroy after duration (default: 5 seconds)
        popup.after(duration, popup.destroy)

    def _fade_status(self, steps=10, interval=50):
        """Fade the status label background color gradually back to neutral."""
        try:
            # Get current background color
            current = self.status_label.cget("background")
            neutral = "#f8f9fa"

            # Convert hex to RGB
            def hex_to_rgb(h): return tuple(int(h[i:i+2], 16) for i in (1, 3, 5))
            def rgb_to_hex(r, g, b): return f"#{r:02x}{g:02x}{b:02x}"

            r1, g1, b1 = hex_to_rgb(current)
            r2, g2, b2 = hex_to_rgb(neutral)

            step_r = (r2 - r1) / steps
            step_g = (g2 - g1) / steps
            step_b = (b2 - b1) / steps

            def fade(i=0):
                if i >= steps:
                    self.status_label.configure(background=neutral, foreground="#333333")
                    self.status_var.set("Ready")
                    return
                r = int(r1 + step_r * i)
                g = int(g1 + step_g * i)
                b = int(b1 + step_b * i)
                self.status_label.configure(background=rgb_to_hex(r, g, b))
                self.root.after(interval, lambda: fade(i + 1))

            fade()
        except Exception:
            # fallback in case of any color conversion errors
            self.status_label.configure(background="#f8f9fa", foreground="#333333")
            self.status_var.set("Ready")

    def _add_hover_effect(self, widget):
        widget.bind("<Enter>", lambda e: widget.configure(cursor="hand2"))
        widget.bind("<Leave>", lambda e: widget.configure(cursor=""))

    def _show_about(self):
        messagebox.showinfo("About", f"{APP_TITLE}\nDynamic Excel-driven data entry with adaptive validation.\nBuilt passionately by tEppy using Python.")

    def _show_help(self):
        """Display help instructions from help.txt in a floating card-like window."""
        help_path = os.path.join(os.path.dirname(__file__), "help.txt")
        if not os.path.exists(help_path):
            messagebox.showinfo("Help File Missing", "No 'help.txt' file found in the app directory.")
            return

        # Read file content
        with open(help_path, "r", encoding="utf-8") as f:
            content = f.read()

        # --- Create floating window ---
        win = tk.Toplevel(self.root)
        win.title("Help & Instructions")
        win.geometry("700x500")
        win.transient(self.root)  # stays above main
        win.resizable(True, True)
        win.configure(bg="#f8f9fa")

        # --- Card Frame (visual styling) ---
        card = ttk.Frame(win, padding=20, relief="raised", borderwidth=2)
        card.pack(expand=True, fill="both", padx=16, pady=16)

        # Scrollable text
        text_frame = ttk.Frame(card)
        text_frame.pack(expand=True, fill="both")

        text_widget = tk.Text(
            text_frame,
            wrap="word",
            font=("Segoe UI", 10),
            relief="flat",
            bg="#ffffff",
            fg="#333333"
        )
        text_widget.insert("1.0", content)
        text_widget.config(state="disabled")

        scroll = ttk.Scrollbar(text_frame, command=text_widget.yview)
        text_widget.config(yscrollcommand=scroll.set)
        scroll.pack(side=tk.RIGHT, fill=tk.Y)
        text_widget.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        # Close button
        close_btn = ttk.Button(card, text="Close", command=win.destroy, style="secondary.TButton")
        close_btn.pack(pady=(10, 0), anchor="e")

        # Subtle shadow effect (optional aesthetic)
        try:
            win.attributes("-alpha", 0.98)
            win.lift()
        except Exception:
            pass

    def on_close(self):
        if self.unsaved_changes:
            res = messagebox.askyesnocancel("Unsaved changes", "You have unsaved changes. Save before exit?")
            if res is None:
                return
            if res:
                if not self.save_file():
                    return

        try:
            self.root.destroy()
        except Exception:
            os._exit(0)
            
# -------------------------
# Application Entry Point
# -------------------------
def main():
    app_root = Window(title=APP_TITLE, themename="cosmo")
    app = DynamicExcelApp(app_root)
    app_root.mainloop()


if __name__ == "__main__":
    main()